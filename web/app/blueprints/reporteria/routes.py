from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from decimal import Decimal

import sqlalchemy as sa
from sqlalchemy.orm import joinedload
from flask import flash, make_response, redirect, render_template, request, url_for
from flask_login import current_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ...auth.decorators import role_required
from ...extensions import db
from ...models import Cargo, Contrato, Empleador, Obra, Trabajador
from ...utils import parse_date
from . import bp


# ============================================================
# Helpers generales
# ============================================================

ESTADOS_RAPIDOS = {
    "vigentes_hoy": "Contratos vigentes a la fecha",
    "vencidos": "Contratos vencidos/terminados",
    "por_vencer_30": "Contratos por vencer en 30 días",
    "por_vencer_60": "Contratos por vencer en 60 días",
    "indefinidos": "Contratos indefinidos",
    "iniciados_30": "Contratos iniciados en últimos 30 días",
}


def _today() -> date:
    return date.today()


def _fmt_date(value) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    return value.strftime("%d-%m-%Y")


def _fmt_money(value) -> str:
    if value is None:
        return ""
    try:
        return f"${int(Decimal(value)):,.0f}".replace(",", ".")
    except Exception:
        return str(value)


def _rut_trabajador(t: Trabajador) -> str:
    rut = (getattr(t, "rut", None) or "").strip()
    dv = (getattr(t, "dv", None) or "").strip().upper()
    return f"{rut}-{dv}" if rut and dv else rut


def _nombre_trabajador(t: Trabajador) -> str:
    partes = [t.ap_paterno or "", t.ap_materno or "", t.nombres or ""]
    return " ".join(p.strip() for p in partes if p and p.strip())


def _tipo_norm(tipo: str | None) -> str:
    return (tipo or "Sin tipo").strip().upper() or "Sin tipo"


def _estado_calculado(c: Contrato, hoy: date | None = None) -> str:
    hoy = hoy or _today()
    estado = (c.estado_contrato or "").strip().upper()
    if estado and estado not in {"VIGENTE", "ACTIVO"}:
        return estado.title()
    if c.fecha_termino and c.fecha_termino < hoy:
        return "Vencido"
    return "Vigente"


def _dias_para_vencer(c: Contrato, hoy: date | None = None):
    hoy = hoy or _today()
    if not c.fecha_termino:
        return None
    return (c.fecha_termino - hoy).days


def _antiguedad_referencial(fecha_inicio, fecha_referencia: date | None = None) -> str:
    if not fecha_inicio:
        return ""
    if isinstance(fecha_inicio, datetime):
        fecha_inicio = fecha_inicio.date()
    ref = fecha_referencia or _today()
    if fecha_inicio > ref:
        return "Inicio futuro"

    total_meses = (ref.year - fecha_inicio.year) * 12 + (ref.month - fecha_inicio.month)
    if ref.day < fecha_inicio.day:
        total_meses -= 1
    total_meses = max(total_meses, 0)
    anios, meses = divmod(total_meses, 12)
    dias_extra = max((ref - fecha_inicio).days, 0)

    if anios and meses:
        return f"{anios} año{'s' if anios != 1 else ''}, {meses} mes{'es' if meses != 1 else ''}"
    if anios:
        return f"{anios} año{'s' if anios != 1 else ''}"
    if meses:
        return f"{meses} mes{'es' if meses != 1 else ''}"
    return f"{dias_extra} día{'s' if dias_extra != 1 else ''}"


def _accion_sugerida_control(r: dict) -> str:
    dias = r.get("dias_para_vencer")
    tipo = (r.get("tipo_contrato") or "").upper()
    estado = (r.get("estado_calculado") or "").upper()

    if dias is None:
        if "INDEFIN" in tipo:
            return "Sin acción por vencimiento"
        return "Revisar tipo/fecha de término"
    if dias < 0 or estado != "VIGENTE":
        return "Regularizar estado contractual"
    if dias <= 7:
        return "Definir renovación o término urgente"
    if dias <= 15:
        return "Solicitar definición a jefatura"
    if dias <= 30:
        return "Evaluar renovación/anexo"
    if dias <= 60:
        return "Programar revisión preventiva"
    return "Monitorear"


def _obras_accesibles():
    q = Obra.query
    if current_user.has_role("ADMIN"):
        return q.order_by(Obra.centro_costo.asc().nullslast(), Obra.nombre.asc()).all()

    obra_ids = [o.id for o in (current_user.obras or [])]
    if not obra_ids:
        return []
    return (
        q.filter(Obra.id.in_(obra_ids))
        .order_by(Obra.centro_costo.asc().nullslast(), Obra.nombre.asc())
        .all()
    )


def _aplicar_acceso_obras(query):
    if current_user.has_role("ADMIN"):
        return query
    obra_ids = [o.id for o in (current_user.obras or [])]
    if not obra_ids:
        return query.filter(db.text("1=0"))
    return query.filter(Contrato.obra_id.in_(obra_ids))


def _safe_int_list(name: str) -> list[int]:
    raw_values = request.args.getlist(name)
    values: list[int] = []
    for raw in raw_values:
        try:
            if str(raw).strip():
                values.append(int(raw))
        except Exception:
            continue
    return values


def _filtros_desde_request() -> dict:
    return {
        "q": (request.args.get("q") or "").strip(),
        "obra_ids": _safe_int_list("obra_id"),
        "cargo_ids": _safe_int_list("cargo_id"),
        "empleador_ids": _safe_int_list("empleador_id"),
        "estado": (request.args.get("estado") or "").strip().upper(),
        "estado_rapido": (request.args.get("estado_rapido") or "").strip(),
        "tipo_contrato": (request.args.get("tipo_contrato") or "").strip(),
        "inicio_desde_raw": (request.args.get("inicio_desde") or "").strip(),
        "inicio_hasta_raw": (request.args.get("inicio_hasta") or "").strip(),
        "termino_desde_raw": (request.args.get("termino_desde") or "").strip(),
        "termino_hasta_raw": (request.args.get("termino_hasta") or "").strip(),
    }


def _query_contratos(filtros: dict):
    hoy = _today()
    query = (
        Contrato.query
        .join(Trabajador, Trabajador.id == Contrato.trabajador_id)
        .outerjoin(Obra, Obra.id == Contrato.obra_id)
        .outerjoin(Cargo, Cargo.id == Contrato.cargo_id)
        .outerjoin(Empleador, Empleador.id == Contrato.empleador_id)
        .options(
            joinedload(Contrato.trabajador),
            joinedload(Contrato.obra),
            joinedload(Contrato.cargo),
            joinedload(Contrato.empleador),
        )
    )
    query = _aplicar_acceso_obras(query)

    if filtros["q"]:
        like = f"%{filtros['q']}%"
        query = query.filter(sa.or_(
            Trabajador.nombres.ilike(like),
            Trabajador.ap_paterno.ilike(like),
            Trabajador.ap_materno.ilike(like),
            Trabajador.rut.ilike(like),
            Cargo.nombre.ilike(like),
            Obra.nombre.ilike(like),
            Obra.centro_costo.ilike(like),
            Empleador.razon_social.ilike(like),
        ))

    if filtros["obra_ids"]:
        query = query.filter(Contrato.obra_id.in_(filtros["obra_ids"]))
    if filtros["cargo_ids"]:
        query = query.filter(Contrato.cargo_id.in_(filtros["cargo_ids"]))
    if filtros["empleador_ids"]:
        query = query.filter(Contrato.empleador_id.in_(filtros["empleador_ids"]))

    if filtros["estado"]:
        query = query.filter(sa.func.upper(Contrato.estado_contrato) == filtros["estado"])

    if filtros["tipo_contrato"]:
        query = query.filter(Contrato.tipo_contrato == filtros["tipo_contrato"])

    inicio_desde = parse_date(filtros["inicio_desde_raw"])
    inicio_hasta = parse_date(filtros["inicio_hasta_raw"])
    termino_desde = parse_date(filtros["termino_desde_raw"])
    termino_hasta = parse_date(filtros["termino_hasta_raw"])

    if inicio_desde:
        query = query.filter(Contrato.fecha_inicio >= inicio_desde)
    if inicio_hasta:
        query = query.filter(Contrato.fecha_inicio <= inicio_hasta)
    if termino_desde:
        query = query.filter(Contrato.fecha_termino >= termino_desde)
    if termino_hasta:
        query = query.filter(Contrato.fecha_termino <= termino_hasta)

    estado_rapido = filtros["estado_rapido"]
    if estado_rapido == "vigentes_hoy":
        query = query.filter(
            sa.or_(Contrato.fecha_inicio.is_(None), Contrato.fecha_inicio <= hoy),
            sa.or_(Contrato.fecha_termino.is_(None), Contrato.fecha_termino >= hoy),
            sa.func.upper(sa.func.coalesce(Contrato.estado_contrato, "VIGENTE")).in_(["VIGENTE", "ACTIVO"]),
        )
    elif estado_rapido == "vencidos":
        query = query.filter(sa.or_(
            Contrato.fecha_termino < hoy,
            sa.func.upper(sa.func.coalesce(Contrato.estado_contrato, "")).in_(["TERMINADO", "VENCIDO", "FINIQUITADO"]),
        ))
    elif estado_rapido == "por_vencer_30":
        query = query.filter(Contrato.fecha_termino >= hoy, Contrato.fecha_termino <= hoy + timedelta(days=30))
    elif estado_rapido == "por_vencer_60":
        query = query.filter(Contrato.fecha_termino >= hoy, Contrato.fecha_termino <= hoy + timedelta(days=60))
    elif estado_rapido == "indefinidos":
        query = query.filter(sa.or_(
            Contrato.fecha_termino.is_(None),
            sa.func.upper(sa.func.coalesce(Contrato.tipo_contrato, "")).ilike("%INDEFIN%"),
        ))
    elif estado_rapido == "iniciados_30":
        query = query.filter(Contrato.fecha_inicio >= hoy - timedelta(days=30), Contrato.fecha_inicio <= hoy)

    return query.order_by(
        Obra.nombre.asc().nullslast(),
        Trabajador.ap_paterno.asc(),
        Trabajador.ap_materno.asc().nullslast(),
        Trabajador.nombres.asc(),
        Contrato.fecha_inicio.desc().nullslast(),
    )


def _contratos_to_rows(contratos: list[Contrato]) -> list[dict]:
    hoy = _today()
    rows: list[dict] = []
    for c in contratos:
        t = c.trabajador
        dias = _dias_para_vencer(c, hoy)
        rows.append({
            "contrato_id": c.id,
            "trabajador_id": t.id if t else None,
            "rut": _rut_trabajador(t) if t else "",
            "trabajador": _nombre_trabajador(t) if t else "Sin trabajador",
            "obra": c.obra.nombre if c.obra else "Sin obra",
            "centro_costo": c.obra.centro_costo if c.obra else "",
            "empleador": c.empleador.razon_social if c.empleador else "Sin empleador",
            "cargo": c.cargo.nombre if c.cargo else "Sin cargo",
            "tipo_contrato": c.tipo_contrato or "Sin tipo",
            "estado_contrato": c.estado_contrato or "Sin estado",
            "estado_calculado": _estado_calculado(c, hoy),
            "fecha_inicio": c.fecha_inicio,
            "fecha_termino": c.fecha_termino,
            "antiguedad": _antiguedad_referencial(c.fecha_inicio, hoy),
            "dias_para_vencer": dias,
            "horas_semanales": c.horas_semanales,
            "sueldo_base": c.sueldo_base,
            "tipo_trabajador": t.tipo_trabajador if t else "",
            "telefono": t.telefono if t else "",
            "correo": t.correo if t else "",
        })
    return rows


def _metricas(rows: list[dict]) -> dict:
    hoy = _today()
    trabajadores_unicos = {r["trabajador_id"] for r in rows if r.get("trabajador_id")}
    vigentes = [r for r in rows if r["estado_calculado"] == "Vigente"]
    vencidos = [r for r in rows if r["estado_calculado"] != "Vigente"]
    por_vencer_30 = [r for r in rows if r["dias_para_vencer"] is not None and 0 <= r["dias_para_vencer"] <= 30]
    por_vencer_60 = [r for r in rows if r["dias_para_vencer"] is not None and 0 <= r["dias_para_vencer"] <= 60]
    indefinidos = [r for r in rows if not r["fecha_termino"] or "INDEFIN" in (r["tipo_contrato"] or "").upper()]
    iniciados_30 = [r for r in rows if r["fecha_inicio"] and hoy - timedelta(days=30) <= r["fecha_inicio"] <= hoy]
    directos = [r for r in rows if _tipo_norm(r.get("tipo_trabajador")) == "DIRECTO"]
    return {
        "total_contratos": len(rows),
        "trabajadores_unicos": len(trabajadores_unicos),
        "vigentes": len(vigentes),
        "vencidos": len(vencidos),
        "por_vencer_30": len(por_vencer_30),
        "por_vencer_60": len(por_vencer_60),
        "indefinidos": len(indefinidos),
        "iniciados_30": len(iniciados_30),
        "directos": len(directos),
        "indirectos": max(len(rows) - len(directos), 0),
        "fecha_reporte": hoy,
    }


def _catalogos():
    obras = _obras_accesibles()
    obra_ids = [o.id for o in obras]

    cargos = Cargo.query.order_by(Cargo.nombre.asc()).all()
    empleadores = Empleador.query.order_by(Empleador.razon_social.asc()).all()

    tipo_rows = (
        db.session.query(Contrato.tipo_contrato)
        .filter(Contrato.tipo_contrato.isnot(None))
        .distinct()
        .order_by(Contrato.tipo_contrato.asc())
        .all()
    )
    tipos_contrato = [r[0] for r in tipo_rows if (r[0] or "").strip()]

    return {
        "obras": obras,
        "obra_ids_accesibles": obra_ids,
        "cargos": cargos,
        "empleadores": empleadores,
        "tipos_contrato": tipos_contrato,
    }


def _build_context(limit_for_screen: bool = True) -> dict:
    filtros = _filtros_desde_request()
    contratos = _query_contratos(filtros).all()
    rows = _contratos_to_rows(contratos)
    metricas = _metricas(rows)
    catalogos = _catalogos()
    return {
        "filtros": filtros,
        "rows": rows[:500] if limit_for_screen else rows,
        "all_rows": rows,
        "total_rows": len(rows),
        "is_limited": limit_for_screen and len(rows) > 500,
        "metricas": metricas,
        "estados_rapidos": ESTADOS_RAPIDOS,
        **catalogos,
    }


def _resumen_ejecutivo(rows: list[dict]) -> dict:
    por_obra = defaultdict(lambda: {
        "obra": "Sin obra",
        "centro_costo": "",
        "contratos": 0,
        "trabajadores": set(),
        "vigentes": 0,
        "vencidos": 0,
        "por_vencer_30": 0,
        "indefinidos": 0,
        "sueldo_base_total": Decimal("0"),
    })

    cargo_counter = Counter()
    tipo_counter = Counter()
    empleador_counter = Counter()
    estado_counter = Counter()

    for r in rows:
        key = f"{r['obra']}||{r['centro_costo']}"
        item = por_obra[key]
        item["obra"] = r["obra"]
        item["centro_costo"] = r["centro_costo"]
        item["contratos"] += 1
        if r.get("trabajador_id"):
            item["trabajadores"].add(r["trabajador_id"])
        if r["estado_calculado"] == "Vigente":
            item["vigentes"] += 1
        else:
            item["vencidos"] += 1
        if r["dias_para_vencer"] is not None and 0 <= r["dias_para_vencer"] <= 30:
            item["por_vencer_30"] += 1
        if not r["fecha_termino"] or "INDEFIN" in (r["tipo_contrato"] or "").upper():
            item["indefinidos"] += 1
        try:
            if r["sueldo_base"] is not None:
                item["sueldo_base_total"] += Decimal(r["sueldo_base"])
        except Exception:
            pass

        cargo_counter[r["cargo"] or "Sin cargo"] += 1
        tipo_counter[r["tipo_contrato"] or "Sin tipo"] += 1
        empleador_counter[r["empleador"] or "Sin empleador"] += 1
        estado_counter[r["estado_calculado"] or "Sin estado"] += 1

    obras = []
    for item in por_obra.values():
        item = dict(item)
        item["trabajadores_unicos"] = len(item.pop("trabajadores"))
        obras.append(item)
    obras.sort(key=lambda x: (-x["vigentes"], x["obra"]))

    vencimientos_60 = sorted(
        [r for r in rows if r["dias_para_vencer"] is not None and 0 <= r["dias_para_vencer"] <= 60],
        key=lambda r: (r["dias_para_vencer"], r["trabajador"]),
    )
    iniciados_30 = sorted(
        [r for r in rows if r["fecha_inicio"] and _today() - timedelta(days=30) <= r["fecha_inicio"] <= _today()],
        key=lambda r: (r["fecha_inicio"] or date.min, r["trabajador"]),
        reverse=True,
    )

    return {
        "obras": obras,
        "cargos_top": cargo_counter.most_common(15),
        "tipos_contrato": tipo_counter.most_common(),
        "empleadores": empleador_counter.most_common(),
        "estados": estado_counter.most_common(),
        "vencimientos_60": vencimientos_60[:100],
        "iniciados_30": iniciados_30[:100],
    }


def _excel_title(ws, title: str, end_col: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor="1F2937")
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")


def _style_headers(ws, row: int, end_col: int):
    fill = PatternFill("solid", fgColor="E5E7EB")
    font = Font(bold=True, color="111827")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_borders(ws):
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


# ============================================================
# Pantallas
# ============================================================

@bp.get("/")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def index():
    ctx = _build_context(limit_for_screen=True)
    return render_template("reporteria/index.html", **ctx)


@bp.get("/panel")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def panel():
    ctx = _build_context(limit_for_screen=False)
    ctx["resumen"] = _resumen_ejecutivo(ctx["all_rows"])
    return render_template("reporteria/panel.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)


# ============================================================
# Exportación Excel detalle
# ============================================================

@bp.get("/trabajadores.xlsx")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def trabajadores_xlsx():
    ctx = _build_context(limit_for_screen=False)
    rows = ctx["all_rows"]
    metricas = ctx["metricas"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte trabajadores"

    _excel_title(ws, "Reporte de trabajadores y contratos · Sistema Grupo CS", 18)
    ws["A3"] = "Fecha reporte"
    ws["B3"] = _fmt_date(metricas["fecha_reporte"])
    ws["D3"] = "Contratos"
    ws["E3"] = metricas["total_contratos"]
    ws["G3"] = "Trabajadores únicos"
    ws["H3"] = metricas["trabajadores_unicos"]
    ws["J3"] = "Vigentes"
    ws["K3"] = metricas["vigentes"]
    ws["M3"] = "Por vencer 30 días"
    ws["N3"] = metricas["por_vencer_30"]

    headers = [
        "Contrato ID", "RUT", "Trabajador", "Empleador", "Obra", "Centro costo",
        "Cargo", "Tipo contrato", "Estado", "Estado calculado", "Inicio", "Término",
        "Días vencimiento", "Horas sem.", "Sueldo base", "Tipo trabajador", "Teléfono", "Correo",
    ]
    start_row = 5
    for col, header in enumerate(headers, start=1):
        ws.cell(start_row, col, header)
    _style_headers(ws, start_row, len(headers))

    for r_idx, r in enumerate(rows, start=start_row + 1):
        values = [
            r["contrato_id"], r["rut"], r["trabajador"], r["empleador"], r["obra"], r["centro_costo"],
            r["cargo"], r["tipo_contrato"], r["estado_contrato"], r["estado_calculado"],
            _fmt_date(r["fecha_inicio"]), _fmt_date(r["fecha_termino"]), r["dias_para_vencer"],
            r["horas_semanales"], int(r["sueldo_base"] or 0) if r["sueldo_base"] is not None else None,
            r["tipo_trabajador"], r["telefono"], r["correo"],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(r_idx, col, value)
            if col == 15 and value is not None:
                cell.number_format = '$ #,##0'
    _apply_borders(ws)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{start_row}:R{max(start_row, start_row + len(rows))}"

    widths = [12, 14, 34, 32, 28, 18, 30, 18, 16, 18, 13, 13, 14, 11, 14, 18, 16, 32]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"reporte_trabajadores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = make_response(bio.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# ============================================================
# Exportación Excel resumen ejecutivo
# ============================================================

@bp.get("/resumen.xlsx")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def resumen_xlsx():
    ctx = _build_context(limit_for_screen=False)
    resumen = _resumen_ejecutivo(ctx["all_rows"])
    metricas = ctx["metricas"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen por obra"
    _excel_title(ws, "Resumen ejecutivo por obra · Sistema Grupo CS", 9)
    ws["A3"] = "Fecha reporte"
    ws["B3"] = _fmt_date(metricas["fecha_reporte"])
    headers = ["Obra", "Centro costo", "Contratos", "Trabajadores únicos", "Vigentes", "Vencidos", "Por vencer 30", "Indefinidos", "Sueldo base total"]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    _style_headers(ws, 5, len(headers))
    for idx, item in enumerate(resumen["obras"], start=6):
        ws.cell(idx, 1, item["obra"])
        ws.cell(idx, 2, item["centro_costo"])
        ws.cell(idx, 3, item["contratos"])
        ws.cell(idx, 4, item["trabajadores_unicos"])
        ws.cell(idx, 5, item["vigentes"])
        ws.cell(idx, 6, item["vencidos"])
        ws.cell(idx, 7, item["por_vencer_30"])
        ws.cell(idx, 8, item["indefinidos"])
        ws.cell(idx, 9, int(item["sueldo_base_total"] or 0))
        ws.cell(idx, 9).number_format = '$ #,##0'
    ws.auto_filter.ref = f"A5:I{max(5, 5 + len(resumen['obras']))}"
    ws.freeze_panes = "A6"
    for idx, width in enumerate([34, 18, 12, 18, 12, 12, 16, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_borders(ws)

    ws2 = wb.create_sheet("Top cargos")
    _excel_title(ws2, "Dotación por cargo", 2)
    ws2.append([]); ws2.append([]); ws2.append([])
    ws2.append(["Cargo", "Contratos"])
    _style_headers(ws2, 5, 2)
    for cargo, total in resumen["cargos_top"]:
        ws2.append([cargo, total])
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 14
    _apply_borders(ws2)

    ws3 = wb.create_sheet("Vencimientos 60 días")
    _excel_title(ws3, "Contratos por vencer en los próximos 60 días", 8)
    ws3.append([]); ws3.append([]); ws3.append([])
    venc_headers = ["Días", "Inicio", "Término", "Antigüedad", "Trabajador", "RUT", "Obra", "Cargo", "Tipo", "Estado"]
    ws3.append(venc_headers)
    _style_headers(ws3, 5, len(venc_headers))
    for r in resumen["vencimientos_60"]:
        ws3.append([r["dias_para_vencer"], _fmt_date(r["fecha_inicio"]), _fmt_date(r["fecha_termino"]), r.get("antiguedad"), r["trabajador"], r["rut"], r["obra"], r["cargo"], r["tipo_contrato"], r["estado_calculado"]])
    for idx, width in enumerate([10, 13, 13, 16, 34, 14, 30, 34, 18, 16], start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = width
    _apply_borders(ws3)

    ws4 = wb.create_sheet("Distribuciones")
    _excel_title(ws4, "Distribuciones generales", 4)
    row = 4
    for titulo, data in (("Tipos de contrato", resumen["tipos_contrato"]), ("Empleadores", resumen["empleadores"]), ("Estados calculados", resumen["estados"])):
        ws4.cell(row, 1, titulo).font = Font(bold=True)
        row += 1
        ws4.cell(row, 1, "Categoría")
        ws4.cell(row, 2, "Contratos")
        _style_headers(ws4, row, 2)
        row += 1
        for label, total in data:
            ws4.cell(row, 1, label)
            ws4.cell(row, 2, total)
            row += 1
        row += 2
    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 14
    _apply_borders(ws4)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"resumen_reporteria_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = make_response(bio.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# ============================================================
# Exportación PDF
# ============================================================

@bp.get("/trabajadores.pdf")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def trabajadores_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        flash(
            "Motor PDF no disponible (falta WeasyPrint o dependencias). Instala WeasyPrint para habilitar la exportación.",
            "error",
        )
        return redirect(url_for("reporteria.index", **request.args))

    ctx = _build_context(limit_for_screen=False)
    html = render_template("reporteria/trabajadores_pdf.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)
    pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()

    filename = f"reporte_trabajadores_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    return response


@bp.get("/resumen.pdf")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def resumen_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        flash(
            "Motor PDF no disponible (falta WeasyPrint o dependencias). Instala WeasyPrint para habilitar la exportación.",
            "error",
        )
        return redirect(url_for("reporteria.panel", **request.args))

    ctx = _build_context(limit_for_screen=False)
    ctx["resumen"] = _resumen_ejecutivo(ctx["all_rows"])
    html = render_template("reporteria/resumen_pdf.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)
    pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()

    filename = f"resumen_reporteria_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    return response

# ============================================================
# Reportería v3 · Control contractual y calidad de datos
# ============================================================

CONTROL_BUCKETS = [
    ("vencen_7", "Vencen en 7 días", 0, 7, "danger"),
    ("vencen_15", "Vencen en 15 días", 8, 15, "warning"),
    ("vencen_30", "Vencen en 30 días", 16, 30, "info"),
    ("vencen_60", "Vencen en 60 días", 31, 60, "secondary"),
]


def _control_contractual(rows: list[dict]) -> dict:
    hoy = _today()
    buckets = {
        key: {"key": key, "label": label, "min": dmin, "max": dmax, "color": color, "rows": []}
        for key, label, dmin, dmax, color in CONTROL_BUCKETS
    }
    vencidos = []
    indefinidos = []
    sin_fecha_termino_no_indef = []

    for r in rows:
        dias = r.get("dias_para_vencer")
        tipo = (r.get("tipo_contrato") or "").upper()
        estado_calc = r.get("estado_calculado") or ""
        estado_original = (r.get("estado_contrato") or "").upper()

        if dias is not None:
            if dias < 0 or estado_calc != "Vigente" or estado_original in {"TERMINADO", "VENCIDO", "FINIQUITADO"}:
                vencidos.append(r)
            else:
                for key, _label, dmin, dmax, _color in CONTROL_BUCKETS:
                    if dmin <= dias <= dmax:
                        buckets[key]["rows"].append(r)
                        break
        else:
            if "INDEFIN" in tipo:
                indefinidos.append(r)
            else:
                sin_fecha_termino_no_indef.append(r)

    for item in buckets.values():
        item["rows"].sort(key=lambda x: (x.get("dias_para_vencer") or 9999, x.get("trabajador") or ""))
    vencidos.sort(key=lambda x: ((x.get("fecha_termino") or date.min), x.get("trabajador") or ""), reverse=True)
    sin_fecha_termino_no_indef.sort(key=lambda x: (x.get("obra") or "", x.get("trabajador") or ""))
    indefinidos.sort(key=lambda x: (x.get("obra") or "", x.get("trabajador") or ""))

    return {
        "fecha_reporte": hoy,
        "buckets": list(buckets.values()),
        "vencidos": vencidos,
        "indefinidos": indefinidos,
        "sin_fecha_termino_no_indef": sin_fecha_termino_no_indef,
        "total_alertas": sum(len(b["rows"]) for b in buckets.values()) + len(vencidos) + len(sin_fecha_termino_no_indef),
        "total_7": len(buckets["vencen_7"]["rows"]),
        "total_15": len(buckets["vencen_15"]["rows"]),
        "total_30": len(buckets["vencen_30"]["rows"]),
        "total_60": len(buckets["vencen_60"]["rows"]),
        "total_vencidos": len(vencidos),
        "total_indefinidos": len(indefinidos),
        "total_sin_fecha_termino_no_indef": len(sin_fecha_termino_no_indef),
    }


def _auditoria_datos(rows: list[dict]) -> dict:
    issues: list[dict] = []
    hoy = _today()

    def add(r: dict, severidad: str, categoria: str, problema: str, sugerencia: str):
        issues.append({
            "severidad": severidad,
            "categoria": categoria,
            "problema": problema,
            "sugerencia": sugerencia,
            "contrato_id": r.get("contrato_id"),
            "rut": r.get("rut") or "",
            "trabajador": r.get("trabajador") or "Sin trabajador",
            "obra": r.get("obra") or "Sin obra",
            "cargo": r.get("cargo") or "Sin cargo",
            "empleador": r.get("empleador") or "Sin empleador",
            "tipo_contrato": r.get("tipo_contrato") or "Sin tipo",
            "estado_calculado": r.get("estado_calculado") or "Sin estado",
        })

    for r in rows:
        tipo = (r.get("tipo_contrato") or "").upper()
        estado = (r.get("estado_contrato") or "").upper()
        fecha_inicio = r.get("fecha_inicio")
        fecha_termino = r.get("fecha_termino")
        rut = (r.get("rut") or "").strip()
        sueldo = r.get("sueldo_base")

        if not rut or rut == "-":
            add(r, "Alta", "Trabajador", "Trabajador sin RUT completo", "Completar RUT y dígito verificador en la ficha del trabajador.")
        if r.get("obra") == "Sin obra":
            add(r, "Alta", "Contrato", "Contrato sin obra asociada", "Asignar obra para reportes, nóminas y control operativo.")
        if r.get("cargo") == "Sin cargo":
            add(r, "Alta", "Contrato", "Contrato sin cargo asociado", "Asignar cargo correcto para documentos, reportes y descripciones de cargo.")
        if r.get("empleador") == "Sin empleador":
            add(r, "Alta", "Contrato", "Contrato sin empleador asociado", "Asignar empresa empleadora para documentos laborales y trazabilidad.")
        if not fecha_inicio:
            add(r, "Alta", "Contrato", "Contrato sin fecha de inicio", "Registrar fecha de inicio contractual.")
        if fecha_inicio and fecha_termino and fecha_termino < fecha_inicio:
            add(r, "Alta", "Fechas", "Fecha de término anterior a fecha de inicio", "Corregir rango contractual; este dato afecta vigencia y reportería.")
        if "PLAZO" in tipo and not fecha_termino:
            add(r, "Media", "Fechas", "Contrato a plazo sin fecha de término", "Registrar fecha de término o revisar si corresponde contrato indefinido.")
        if fecha_termino and fecha_termino < hoy and estado in {"VIGENTE", "ACTIVO", ""}:
            add(r, "Alta", "Vigencia", "Contrato figura vigente pero su fecha de término ya pasó", "Actualizar estado, renovar, anexar o iniciar proceso de término según corresponda.")
        if not sueldo:
            add(r, "Media", "Remuneración", "Contrato sin sueldo base registrado o con valor cero", "Completar sueldo base contractual para reportes y documentos.")
        if r.get("centro_costo") in {None, ""}:
            add(r, "Baja", "Obra", "Obra sin centro de costo visible en el contrato", "Revisar centro de costo en la obra asociada.")
        if not (r.get("tipo_contrato") or "").strip() or r.get("tipo_contrato") == "Sin tipo":
            add(r, "Media", "Contrato", "Contrato sin tipo de contrato", "Definir tipo: plazo fijo, indefinido, faena u otro según corresponda.")

    orden = {"Alta": 0, "Media": 1, "Baja": 2}
    issues.sort(key=lambda x: (orden.get(x["severidad"], 9), x["categoria"], x["trabajador"]))

    por_severidad = Counter(i["severidad"] for i in issues)
    por_categoria = Counter(i["categoria"] for i in issues)
    contratos_afectados = {i["contrato_id"] for i in issues if i.get("contrato_id")}

    return {
        "fecha_reporte": hoy,
        "issues": issues,
        "total_issues": len(issues),
        "contratos_afectados": len(contratos_afectados),
        "altas": por_severidad.get("Alta", 0),
        "medias": por_severidad.get("Media", 0),
        "bajas": por_severidad.get("Baja", 0),
        "por_categoria": por_categoria.most_common(),
        "por_severidad": por_severidad,
    }


def _append_control_sheet(wb: Workbook, control: dict):
    ws = wb.active
    ws.title = "Control contractual"
    _excel_title(ws, "Control contractual · vencimientos y alertas", 10)
    ws["A3"] = "Fecha reporte"
    ws["B3"] = _fmt_date(control["fecha_reporte"])
    headers = ["Bloque", "Días", "Inicio", "Término", "Antigüedad", "Trabajador", "RUT", "Obra", "Cargo", "Tipo contrato", "Estado", "Acción sugerida", "Contrato ID"]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    _style_headers(ws, 5, len(headers))
    row = 6

    def write_row(bloque: str, r: dict):
        nonlocal row
        values = [bloque, r.get("dias_para_vencer"), _fmt_date(r.get("fecha_inicio")), _fmt_date(r.get("fecha_termino")), r.get("antiguedad"), r.get("trabajador"), r.get("rut"), r.get("obra"), r.get("cargo"), r.get("tipo_contrato"), r.get("estado_calculado"), _accion_sugerida_control(r), r.get("contrato_id")]
        for col, value in enumerate(values, start=1):
            ws.cell(row, col, value)
        row += 1

    for bucket in control["buckets"]:
        for r in bucket["rows"]:
            write_row(bucket["label"], r)
    for r in control["vencidos"]:
        write_row("Vencidos / revisar estado", r)
    for r in control["sin_fecha_termino_no_indef"]:
        write_row("Sin fecha término no indefinido", r)

    ws.auto_filter.ref = f"A5:M{max(5, row - 1)}"
    ws.freeze_panes = "A6"
    for idx, width in enumerate([28, 10, 13, 13, 16, 34, 14, 30, 34, 20, 16, 30, 12], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_borders(ws)


def _append_auditoria_sheet(wb: Workbook, auditoria: dict):
    ws = wb.create_sheet("Auditoría datos")
    _excel_title(ws, "Auditoría de calidad de datos · Sistema Grupo CS", 12)
    ws["A3"] = "Fecha reporte"
    ws["B3"] = _fmt_date(auditoria["fecha_reporte"])
    headers = ["Severidad", "Categoría", "Problema", "Sugerencia", "Contrato ID", "RUT", "Trabajador", "Obra", "Cargo", "Empleador", "Tipo contrato", "Estado"]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    _style_headers(ws, 5, len(headers))
    for idx, item in enumerate(auditoria["issues"], start=6):
        values = [item["severidad"], item["categoria"], item["problema"], item["sugerencia"], item["contrato_id"], item["rut"], item["trabajador"], item["obra"], item["cargo"], item["empleador"], item["tipo_contrato"], item["estado_calculado"]]
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col, value)
    ws.auto_filter.ref = f"A5:L{max(5, 5 + len(auditoria['issues']))}"
    ws.freeze_panes = "A6"
    for idx, width in enumerate([13, 16, 36, 48, 12, 14, 34, 30, 30, 34, 20, 16], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_borders(ws)


@bp.get("/control-contractual")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def control_contractual():
    ctx = _build_context(limit_for_screen=False)
    ctx["control"] = _control_contractual(ctx["all_rows"])
    ctx["auditoria"] = _auditoria_datos(ctx["all_rows"])
    return render_template("reporteria/control_contractual.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)


@bp.get("/auditoria-datos")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def auditoria_datos():
    ctx = _build_context(limit_for_screen=False)
    ctx["auditoria"] = _auditoria_datos(ctx["all_rows"])
    return render_template("reporteria/auditoria_datos.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)


@bp.get("/control-contractual.xlsx")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def control_contractual_xlsx():
    ctx = _build_context(limit_for_screen=False)
    control = _control_contractual(ctx["all_rows"])
    auditoria = _auditoria_datos(ctx["all_rows"])
    wb = Workbook()
    _append_control_sheet(wb, control)
    _append_auditoria_sheet(wb, auditoria)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"control_contractual_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = make_response(bio.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@bp.get("/auditoria-datos.xlsx")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def auditoria_datos_xlsx():
    ctx = _build_context(limit_for_screen=False)
    auditoria = _auditoria_datos(ctx["all_rows"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _excel_title(ws, "Resumen auditoría de datos · Sistema Grupo CS", 5)
    ws["A3"] = "Hallazgos"
    ws["B3"] = auditoria["total_issues"]
    ws["C3"] = "Contratos afectados"
    ws["D3"] = auditoria["contratos_afectados"]
    _append_auditoria_sheet(wb, auditoria)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"auditoria_datos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = make_response(bio.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def _seguimiento_contractual_rows(rows: list[dict]) -> list[dict]:
    control = _control_contractual(rows)
    seguimiento: list[dict] = []
    for bucket in control["buckets"]:
        for r in bucket["rows"]:
            item = dict(r)
            item["bloque"] = bucket["label"]
            item["accion_sugerida"] = _accion_sugerida_control(r)
            item["estado_gestion"] = "Pendiente de revisión"
            item["responsable"] = ""
            item["observacion_gestion"] = ""
            seguimiento.append(item)
    for r in control["vencidos"]:
        item = dict(r)
        item["bloque"] = "Vencidos / revisar estado"
        item["accion_sugerida"] = _accion_sugerida_control(r)
        item["estado_gestion"] = "Urgente"
        item["responsable"] = ""
        item["observacion_gestion"] = ""
        seguimiento.append(item)
    for r in control["sin_fecha_termino_no_indef"]:
        item = dict(r)
        item["bloque"] = "Sin fecha término no indefinido"
        item["accion_sugerida"] = _accion_sugerida_control(r)
        item["estado_gestion"] = "Corregir dato"
        item["responsable"] = ""
        item["observacion_gestion"] = ""
        seguimiento.append(item)

    seguimiento.sort(key=lambda x: (
        9999 if x.get("dias_para_vencer") is None else x.get("dias_para_vencer"),
        x.get("fecha_termino") or date.max,
        x.get("trabajador") or "",
    ))
    return seguimiento


def _append_seguimiento_sheet(wb: Workbook, seguimiento: list[dict]):
    ws = wb.active
    ws.title = "Seguimiento contractual"
    _excel_title(ws, "Seguimiento de gestión contractual · Sistema Grupo CS", 16)
    ws["A3"] = "Fecha reporte"
    ws["B3"] = _fmt_date(_today())
    ws["D3"] = "Registros"
    ws["E3"] = len(seguimiento)
    headers = [
        "Bloque", "Estado gestión", "Acción sugerida", "Responsable", "Observación gestión",
        "Días", "Inicio", "Término", "Antigüedad", "Trabajador", "RUT", "Obra",
        "Cargo", "Tipo contrato", "Estado", "Contrato ID",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(5, col, header)
    _style_headers(ws, 5, len(headers))
    for idx, r in enumerate(seguimiento, start=6):
        values = [
            r.get("bloque"), r.get("estado_gestion"), r.get("accion_sugerida"), r.get("responsable"), r.get("observacion_gestion"),
            r.get("dias_para_vencer"), _fmt_date(r.get("fecha_inicio")), _fmt_date(r.get("fecha_termino")), r.get("antiguedad"),
            r.get("trabajador"), r.get("rut"), r.get("obra"), r.get("cargo"), r.get("tipo_contrato"), r.get("estado_calculado"), r.get("contrato_id"),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(idx, col, value)
    ws.auto_filter.ref = f"A5:P{max(5, 5 + len(seguimiento))}"
    ws.freeze_panes = "A6"
    widths = [28, 22, 30, 20, 38, 10, 13, 13, 16, 34, 14, 30, 34, 20, 16, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    _apply_borders(ws)


@bp.get("/seguimiento-contractual")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def seguimiento_contractual():
    ctx = _build_context(limit_for_screen=False)
    ctx["seguimiento_rows"] = _seguimiento_contractual_rows(ctx["all_rows"])
    ctx["control"] = _control_contractual(ctx["all_rows"])
    return render_template("reporteria/seguimiento_contractual.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)


@bp.get("/seguimiento-contractual.xlsx")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def seguimiento_contractual_xlsx():
    ctx = _build_context(limit_for_screen=False)
    seguimiento = _seguimiento_contractual_rows(ctx["all_rows"])
    wb = Workbook()
    _append_seguimiento_sheet(wb, seguimiento)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"seguimiento_contractual_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = make_response(bio.getvalue())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@bp.get("/seguimiento-contractual.pdf")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def seguimiento_contractual_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        flash("Motor PDF no disponible (falta WeasyPrint o dependencias).", "error")
        return redirect(url_for("reporteria.seguimiento_contractual", **request.args))
    ctx = _build_context(limit_for_screen=False)
    ctx["seguimiento_rows"] = _seguimiento_contractual_rows(ctx["all_rows"])
    html = render_template("reporteria/seguimiento_contractual_pdf.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)
    pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
    filename = f"seguimiento_contractual_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    return response


@bp.get("/control-contractual.pdf")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def control_contractual_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        flash("Motor PDF no disponible (falta WeasyPrint o dependencias).", "error")
        return redirect(url_for("reporteria.control_contractual", **request.args))
    ctx = _build_context(limit_for_screen=False)
    ctx["control"] = _control_contractual(ctx["all_rows"])
    ctx["auditoria"] = _auditoria_datos(ctx["all_rows"])
    html = render_template("reporteria/control_contractual_pdf.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)
    pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
    filename = f"control_contractual_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    return response


@bp.get("/auditoria-datos.pdf")
@role_required("ADMIN", "OPERADOR", "REVISOR")
def auditoria_datos_pdf():
    try:
        from weasyprint import HTML  # type: ignore
    except Exception:
        flash("Motor PDF no disponible (falta WeasyPrint o dependencias).", "error")
        return redirect(url_for("reporteria.auditoria_datos", **request.args))
    ctx = _build_context(limit_for_screen=False)
    ctx["auditoria"] = _auditoria_datos(ctx["all_rows"])
    html = render_template("reporteria/auditoria_datos_pdf.html", **ctx, fmt_date=_fmt_date, fmt_money=_fmt_money)
    pdf_bytes = HTML(string=html, base_url=request.host_url).write_pdf()
    filename = f"auditoria_datos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response = make_response(pdf_bytes)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={filename}"
    return response
