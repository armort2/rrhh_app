# web/app/blueprints/terceros/routes.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from flask import current_app, flash, redirect, render_template, request, url_for
from sqlalchemy import or_

from ...auth.decorators import role_required
from ...extensions import db
from ...models import Banco, Tercero


# ==========================
# Helpers
# ==========================

def _data_dir() -> Path:
    # /app/app -> parent = /app, parent.parent = /
    # En tu proyecto, normalmente quedará /app/data dentro del contenedor.
    # Si tienes volumen montado a /srv/rrhh_app/data => en el contenedor debería ser /app/data.
    return Path("/app/data")


def _safe_str(x: Any) -> str:
    return ("" if x is None else str(x)).strip()


def _as_int_str_or_none(val: Any) -> str | None:
    """
    Convierte valores típicos Excel a string de dígitos.
    """
    if val is None:
        return None

    s = _safe_str(val)
    if not s:
        return None

    # Si viene como número (int/float) directo
    if isinstance(val, (int,)):
        return str(val)
    if isinstance(val, float):
        return str(int(val))

    # Caso "12345.0"
    if s.endswith(".0"):
        s = s[:-2]

    # Intenta parsear numérico general
    try:
        f = float(s)
        return str(int(f))
    except Exception:
        return s


def _parse_rut_parts_from_excel(rut_num: Any, dv: Any) -> tuple[int | None, str | None]:
    r = _as_int_str_or_none(rut_num)
    d = _safe_str(dv).upper()
    if not r or not d:
        return None, None
    try:
        return int(r), d
    except Exception:
        return None, None


def _parse_rut_cta(val: Any) -> tuple[int | None, str | None]:
    """
    Excel trae 'RUT (Cta. Bancaria)' como número (sin guion ni DV separado).
    Regla: último char es DV, resto es rut_num.
    """
    raw = _as_int_str_or_none(val)
    if not raw:
        return None, None
    raw = raw.strip()
    if len(raw) < 2:
        return None, None
    rut_num_str = raw[:-1]
    dv = raw[-1].upper()
    try:
        return int(rut_num_str), dv
    except Exception:
        return None, None


def _read_excel_rows(filepath: Path) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Lee primera hoja de un xlsx y devuelve:
    - headers (fila 1)
    - rows como dict header->value
    """
    wb = load_workbook(filename=str(filepath), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    headers = [(_safe_str(h) if h is not None else "") for h in header_row]
    headers = [h for h in headers if h]  # quita vacíos

    data: list[dict[str, Any]] = []
    for r in rows_iter:
        # mapear solo hasta largo de headers
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else None
        data.append(d)

    return headers, data


@dataclass
class ImportResult:
    inserted: int = 0
    updated: int = 0
    errors: list[str] = None  # type: ignore


def _import_excel(filepath: Path, modo: str = "UPSERT") -> ImportResult:
    """
    modo:
      - UPSERT: inserta nuevos y actualiza existentes por rut_num
      - SOLO_NUEVOS: inserta solo si no existe
    """
    res = ImportResult(inserted=0, updated=0, errors=[])

    expected_cols = {
        "ID Trab.",
        "D. Verif.",
        "Ap. Paterno",
        "Ap. Materno",
        "Nombres",
        "Correo electrónico",
        "RUT (Cta. Bancaria)",
        "Nro. Cuenta Bancaria",
        "Cod. Banco",
    }

    headers, rows = _read_excel_rows(filepath)
    if not headers:
        res.errors.append("El archivo Excel no tiene encabezados o está vacío.")
        return res

    missing = sorted(list(expected_cols - set(headers)))
    if missing:
        res.errors.append(f"Faltan columnas requeridas: {', '.join(missing)}")
        return res

    for idx, row in enumerate(rows, start=2):  # fila 2 en adelante (fila 1 = headers)
        rut_num, dv = _parse_rut_parts_from_excel(row.get("ID Trab."), row.get("D. Verif."))
        if not rut_num or not dv:
            res.errors.append(f"Fila {idx}: RUT/DV inválido (ID Trab./D. Verif.).")
            continue

        ap_paterno = _safe_str(row.get("Ap. Paterno"))
        if not ap_paterno:
            res.errors.append(f"Fila {idx}: 'Ap. Paterno' (o razón social) vacío.")
            continue

        ap_materno = _safe_str(row.get("Ap. Materno")) or None
        nombres = _safe_str(row.get("Nombres")) or None
        correo = _safe_str(row.get("Correo electrónico")) or None

        rut_cta_num, rut_cta_dv = _parse_rut_cta(row.get("RUT (Cta. Bancaria)"))

        cuenta_numero = _as_int_str_or_none(row.get("Nro. Cuenta Bancaria"))
        cuenta_numero = cuenta_numero.strip() if cuenta_numero else None

        cod_banco = _as_int_str_or_none(row.get("Cod. Banco"))
        cod_banco = cod_banco.strip() if cod_banco else None

        banco_id = None
        if cod_banco:
            banco = Banco.query.filter(Banco.codigo_sbif == cod_banco).first()
            if not banco:
                res.errors.append(f"Fila {idx}: Código SBIF banco no encontrado: {cod_banco}")
            else:
                banco_id = banco.id

        tercero = Tercero.query.filter_by(rut_num=rut_num).first()
        if tercero and modo == "SOLO_NUEVOS":
            continue

        if not tercero:
            tercero = Tercero(
                rut_num=rut_num,
                dv=dv,
                ap_paterno=ap_paterno,
                ap_materno=ap_materno,
                nombres=nombres,
                correo=correo,
                banco_id=banco_id,
                cuenta_numero=cuenta_numero,
                rut_cta_num=rut_cta_num,
                rut_cta_dv=rut_cta_dv,
                estado="VIGENTE",
            )
            db.session.add(tercero)
            res.inserted += 1
        else:
            tercero.dv = dv
            tercero.ap_paterno = ap_paterno
            tercero.ap_materno = ap_materno
            tercero.nombres = nombres
            tercero.correo = correo
            tercero.banco_id = banco_id
            tercero.cuenta_numero = cuenta_numero
            tercero.rut_cta_num = rut_cta_num
            tercero.rut_cta_dv = rut_cta_dv
            res.updated += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        res.errors.append(f"Error al guardar cambios: {e}")

    return res


# ==========================
# Rutas
# ==========================

from . import bp  # noqa: E402


@role_required("ADMIN", "OPERADOR", "REVISOR")
@bp.route("/", methods=["GET"])
def listado():
    q = _safe_str(request.args.get("q"))
    estado = _safe_str(request.args.get("estado")) or "VIGENTE"

    query = Tercero.query

    if estado and estado != "TODOS":
        query = query.filter(Tercero.estado == estado)

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Tercero.ap_paterno.ilike(like),
                Tercero.ap_materno.ilike(like),
                Tercero.nombres.ilike(like),
            )
        )

    terceros = query.order_by(
        Tercero.ap_paterno.asc(),
        Tercero.ap_materno.asc(),
        Tercero.nombres.asc(),
    ).all()

    return render_template("terceros/listado.html", terceros=terceros, q=q, estado=estado)


@role_required("ADMIN", "OPERADOR")
@bp.route("/nuevo", methods=["GET", "POST"])
def nuevo():
    bancos = Banco.query.order_by(Banco.nombre.asc()).all()

    if request.method == "POST":
        rut_num = _as_int_str_or_none(request.form.get("rut_num"))
        dv = _safe_str(request.form.get("dv")).upper()
        ap_paterno = _safe_str(request.form.get("ap_paterno"))
        ap_materno = _safe_str(request.form.get("ap_materno")) or None
        nombres = _safe_str(request.form.get("nombres")) or None
        correo = _safe_str(request.form.get("correo")) or None

        estado = _safe_str(request.form.get("estado")) or "VIGENTE"
        banco_id = request.form.get("banco_id") or None
        cuenta_numero = _safe_str(request.form.get("cuenta_numero")) or None

        rut_cta_num = _as_int_str_or_none(request.form.get("rut_cta_num"))
        rut_cta_dv = _safe_str(request.form.get("rut_cta_dv")).upper() or None

        if not rut_num or not dv or not ap_paterno:
            flash("Debe completar al menos RUT, DV y Ap. Paterno/Razón Social.", "warning")
            return render_template("terceros/form.html", tercero=None, bancos=bancos)

        try:
            rut_num_int = int(rut_num)
        except Exception:
            flash("RUT inválido.", "warning")
            return render_template("terceros/form.html", tercero=None, bancos=bancos)

        exists = Tercero.query.filter_by(rut_num=rut_num_int).first()
        if exists:
            flash("Ya existe un tercero con ese RUT.", "warning")
            return render_template("terceros/form.html", tercero=None, bancos=bancos)

        t = Tercero(
            rut_num=rut_num_int,
            dv=dv,
            ap_paterno=ap_paterno,
            ap_materno=ap_materno,
            nombres=nombres,
            correo=correo,
            estado=estado,
            banco_id=int(banco_id) if banco_id else None,
            cuenta_numero=cuenta_numero,
            rut_cta_num=int(rut_cta_num) if rut_cta_num else None,
            rut_cta_dv=rut_cta_dv,
        )

        db.session.add(t)
        db.session.commit()
        flash("Tercero creado correctamente.", "success")
        return redirect(url_for("terceros.listado"))

    return render_template("terceros/form.html", tercero=None, bancos=bancos)


@role_required("ADMIN", "OPERADOR")
@bp.route("/<int:tercero_id>/editar", methods=["GET", "POST"])
def editar(tercero_id: int):
    tercero = Tercero.query.get_or_404(tercero_id)
    bancos = Banco.query.order_by(Banco.nombre.asc()).all()

    if request.method == "POST":
        dv = _safe_str(request.form.get("dv")).upper()
        ap_paterno = _safe_str(request.form.get("ap_paterno"))
        ap_materno = _safe_str(request.form.get("ap_materno")) or None
        nombres = _safe_str(request.form.get("nombres")) or None
        correo = _safe_str(request.form.get("correo")) or None

        estado = _safe_str(request.form.get("estado")) or "VIGENTE"
        banco_id = request.form.get("banco_id") or None
        cuenta_numero = _safe_str(request.form.get("cuenta_numero")) or None

        rut_cta_num = _as_int_str_or_none(request.form.get("rut_cta_num"))
        rut_cta_dv = _safe_str(request.form.get("rut_cta_dv")).upper() or None

        if not dv or not ap_paterno:
            flash("Debe completar DV y Ap. Paterno/Razón Social.", "warning")
            return render_template("terceros/form.html", tercero=tercero, bancos=bancos)

        tercero.dv = dv
        tercero.ap_paterno = ap_paterno
        tercero.ap_materno = ap_materno
        tercero.nombres = nombres
        tercero.correo = correo
        tercero.estado = estado
        tercero.banco_id = int(banco_id) if banco_id else None
        tercero.cuenta_numero = cuenta_numero
        tercero.rut_cta_num = int(rut_cta_num) if rut_cta_num else None
        tercero.rut_cta_dv = rut_cta_dv

        db.session.commit()
        flash("Tercero actualizado correctamente.", "success")
        return redirect(url_for("terceros.listado"))

    return render_template("terceros/form.html", tercero=tercero, bancos=bancos)


@role_required("ADMIN", "OPERADOR")
@bp.route("/importar", methods=["GET", "POST"])
def importar():
    data_dir = _data_dir()
    data_dir.mkdir(parents=True, exist_ok=True)

    excels = sorted([p.name for p in data_dir.glob("*.xlsx")])

    if request.method == "POST":
        modo = _safe_str(request.form.get("modo")) or "UPSERT"

        file = request.files.get("archivo")
        chosen = _safe_str(request.form.get("archivo_existente"))

        target_path: Path | None = None

        if file and file.filename:
            safe_name = f"importacion_terceros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            target_path = data_dir / safe_name
            file.save(target_path)
            flash(f"Archivo subido correctamente: {safe_name}", "success")
        elif chosen:
            target_path = data_dir / chosen
        else:
            flash("Debe subir un archivo o seleccionar uno existente.", "warning")
            return render_template("terceros/importar.html", excels=excels)

        if not target_path.exists():
            flash("Archivo no encontrado en /data.", "danger")
            return render_template("terceros/importar.html", excels=excels)

        res = _import_excel(target_path, modo=modo)

        if res.errors:
            flash(
                f"Importación finalizada con observaciones. Insertados: {res.inserted}, Actualizados: {res.updated}, Errores: {len(res.errors)}",
                "warning",
            )
            return render_template(
                "terceros/importar.html",
                excels=sorted([p.name for p in data_dir.glob("*.xlsx")]),
                resultado=res,
                archivo_usado=target_path.name,
                modo=modo,
            )

        flash(f"Importación OK. Insertados: {res.inserted}, Actualizados: {res.updated}", "success")
        return redirect(url_for("terceros.listado"))

    return render_template("terceros/importar.html", excels=excels)