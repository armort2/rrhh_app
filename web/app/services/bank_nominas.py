# web/app/services/bank_nominas.py
from __future__ import annotations

import os
import re
import uuid
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

# ============================================================
# Plantillas Banco Santander (XLSX maestras en static/xlsx_templates)
# ============================================================
TEMPLATE_PAGOS_REMUN = "santander_pagos_masivos_remuneraciones.xlsx"
TEMPLATE_TRANSFER_MAS = "santander_transferencias_masivas.xlsx"

# ✅ NUEVO: Pagos masivos proveedores (según estructura que enviaste)
TEMPLATE_PAGOS_PROV = "santander_pagos_masivos_proveedores.xlsx"

# ============================================================
# Helpers de período / glosas
# ============================================================

MESES_ES = [
    "",  # 0 (no usado)
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def periodo_label_es(anio: int, mes: int) -> str:
    """Convierte (anio, mes) en texto para glosas: 'Enero 2026'."""
    mes_txt = MESES_ES[mes] if 1 <= mes <= 12 else f"Mes {mes}"
    return f"{mes_txt} {anio}"


def glosa_anticipo(anio: int, mes: int) -> str:
    return f"Anticipo de sueldo - {periodo_label_es(anio, mes)}"


def glosa_sueldo(anio: int, mes: int) -> str:
    return f"Liquidación de Sueldo - {periodo_label_es(anio, mes)}"


# ============================================================
# Normalización de RUT para cargadores bancarios
# ============================================================

def rut_plain_from_parts(rut: str | None, dv: str | None) -> str | None:
    """
    Retorna RUT para banco: sin puntos, sin guión, con DV, DV en mayúscula.
    Ej: rut="11.399.975", dv="6" => "113999756"
    """
    if not rut:
        return None
    base = re.sub(r"[^0-9]", "", str(rut))
    if not base:
        return None
    d = (dv or "").strip().upper()
    if not d:
        return base
    return f"{base}{d}"


def rut_plain_from_str(rut_completo: str | None) -> str | None:
    """
    Soporta entradas tipo:
      - 11.399.975-6
      - 11399975-6
      - 113999756
      - 22334254K
    Retorna: 113999756 / 22334254K
    """
    if not rut_completo:
        return None
    s = str(rut_completo).strip().upper().replace(".", "").replace(" ", "")
    if not s:
        return None
    if "-" in s:
        base, dv = s.split("-", 1)
        base = re.sub(r"[^0-9]", "", base)
        dv = re.sub(r"[^0-9K]", "", dv)
        return f"{base}{dv}" if base and dv else None
    return re.sub(r"[^0-9K]", "", s) or None


# ============================================================
# DTOs (filas bancarias) - Remuneraciones (existente)
# ============================================================

@dataclass(frozen=True)
class BankRow:
    rut_beneficiario: str
    nombre_beneficiario: str
    monto: Decimal
    codigo_banco: str
    cuenta_destino: str
    email_beneficiario: str
    # ✅ NUEVO: glosa por fila (opcional)
    glosa: str = ""


@dataclass(frozen=True)
class BuildResult:
    ok_rows: List[BankRow]
    excluidos: List[Dict[str, Any]]


# ============================================================
# DTOs (filas bancarias) - Proveedores / Solicitudes de Fondos (NUEVO)
# ============================================================

@dataclass(frozen=True)
class ProveedorRow:
    item_id: int
    seccion: str
    centro_costo: str

    rut_beneficiario: str
    nombre_beneficiario: str
    codigo_banco: str
    cuenta_destino: str
    email_beneficiario: str

    nro_fact_1: str
    monto: Decimal
    glosa: str


@dataclass(frozen=True)
class ProveedorBuildResult:
    ok_rows: List[ProveedorRow]
    excluidos: List[Dict[str, Any]]  # {item_id, seccion, motivo, rut, nombre}

@dataclass(frozen=True)
class ProveedorPagoRow:
    pagador: str  # "VALENCIA"/"SALEM" (solo referencia)
    tercero_id: int | None

    rut_tercero: str          # RUT proveedor (normalizado)
    rut_beneficiario: str     # RUT cuenta (rut_cta_num+dv)
    nombre_beneficiario: str

    codigo_banco: str
    cuenta_destino: str
    email_beneficiario: str

    facturas: List[Dict[str, Any]]  # [{nro:str, monto:Decimal}, ...] max 11
    monto_total: Decimal
    glosa: str
    fila_orden: int = 1


# ============================================================
# Helpers de limpieza / validación básica
# ============================================================

def _clean_str(v: Any) -> str:
    return (str(v).strip() if v is not None else "")


def _to_decimal(v: Any) -> Decimal:
    try:
        if v is None:
            return Decimal("0")
        if isinstance(v, Decimal):
            return v
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _is_valid_bank_code(code: str) -> bool:
    c = _clean_str(code)
    return bool(c) and bool(re.fullmatch(r"\d{1,10}", c))


def _is_valid_account(acc: str) -> bool:
    a = _clean_str(acc)
    return bool(re.sub(r"\D", "", a))


def _account_plain(acc: str) -> str:
    return re.sub(r"\D", "", _clean_str(acc))


def _email_or_fallback(email: str | None) -> str:
    e = _clean_str(email)
    return e if e else "transferencias@grupo-cs.cl"


# ============================================================
# Construcción de filas bancarias desde Detalles + Trabajador (existente)
# ============================================================

def build_bank_rows(
    detalles_rows: Iterable[Tuple[Any, Any, Any]],  # (Detalle, Trabajador, Banco)
    amount_attr: str = "monto",
) -> BuildResult:
    ok: List[BankRow] = []
    excluidos: List[Dict[str, Any]] = []

    for det, t, banco in detalles_rows:
        if t is None:
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": "Trabajador no encontrado (registro sin relación)",
                    "rut": getattr(det, "rut", None),
                    "nombre": getattr(det, "nombre_completo", None),
                }
            )
            continue

        monto = _to_decimal(getattr(det, amount_attr, None))
        if monto <= 0:
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": f"{amount_attr} <= 0",
                    "rut": getattr(det, "rut", None),
                    "nombre": getattr(det, "nombre_completo", None),
                }
            )
            continue

        if banco is None:
            banco = getattr(t, "banco", None)

        tercero = bool(getattr(t, "pago_tercero_activo", False))

        if tercero:
            rut_b = rut_plain_from_str(getattr(t, "pago_tercero_rut", None))
            nombre_b = _clean_str(getattr(t, "pago_tercero_nombre", None))
            cuenta_b = _account_plain(getattr(t, "pago_tercero_cuenta_numero", None))
            banco_code = _clean_str(getattr(getattr(t, "pago_tercero_banco", None), "codigo_sbif", None))
            email_b = _email_or_fallback(getattr(t, "correo", None))
        else:
            rut_b = rut_plain_from_parts(getattr(t, "rut", None), getattr(t, "dv", None))
            nombre_b = " ".join(
                p for p in [
                    _clean_str(getattr(t, "ap_paterno", None)),
                    _clean_str(getattr(t, "ap_materno", None)),
                    _clean_str(getattr(t, "nombres", None)),
                ] if p
            ).strip()
            cuenta_raw = _clean_str(getattr(t, "cuenta_numero", None)) or _clean_str(getattr(t, "cuenta_rut", None))
            cuenta_b = _account_plain(cuenta_raw)
            banco_code = _clean_str(getattr(banco, "codigo_sbif", None))
            email_b = _email_or_fallback(getattr(t, "correo", None))

        if not rut_b:
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": "RUT beneficiario vacío/invalidable",
                    "nombre": nombre_b or getattr(det, "nombre_completo", None),
                }
            )
            continue

        if not nombre_b:
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": "Nombre beneficiario vacío",
                    "rut_beneficiario": rut_b,
                }
            )
            continue

        if not _is_valid_bank_code(banco_code):
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": f"Código banco inválido: {banco_code or '(vacío)'}",
                    "rut_beneficiario": rut_b,
                    "nombre": nombre_b,
                }
            )
            continue

        if not _is_valid_account(cuenta_b):
            excluidos.append(
                {
                    "detalle_id": getattr(det, "id", None),
                    "trabajador_id": getattr(det, "trabajador_id", None),
                    "motivo": "Cuenta destino inválida/vacía",
                    "rut_beneficiario": rut_b,
                    "nombre": nombre_b,
                }
            )
            continue

        ok.append(
            BankRow(
                rut_beneficiario=rut_b,
                nombre_beneficiario=nombre_b,
                monto=monto,
                codigo_banco=banco_code,
                cuenta_destino=cuenta_b,
                email_beneficiario=email_b,
            )
        )

    return BuildResult(ok_rows=ok, excluidos=excluidos)


def aggregate_rows(rows: List[BankRow]) -> List[BankRow]:
    acc: Dict[Tuple[str, str, str], BankRow] = {}
    sums: Dict[Tuple[str, str, str], Decimal] = {}

    for r in rows:
        k = (r.rut_beneficiario, r.codigo_banco, r.cuenta_destino)
        if k not in acc:
            acc[k] = r
            sums[k] = r.monto
        else:
            sums[k] = sums[k] + r.monto

    out: List[BankRow] = []
    for k, base in acc.items():
        out.append(
            BankRow(
                rut_beneficiario=base.rut_beneficiario,
                nombre_beneficiario=base.nombre_beneficiario,
                monto=sums[k],
                codigo_banco=base.codigo_banco,
                cuenta_destino=base.cuenta_destino,
                email_beneficiario=base.email_beneficiario,
            )
        )
    return out

def _get_first_attr(obj: Any, *names: str) -> Any:
    """Devuelve el primer atributo existente/no-nulo."""
    for n in names:
        if obj is None:
            return None
        if hasattr(obj, n):
            v = getattr(obj, n, None)
            if v is not None and v != "":
                return v
    return None


def _numero_documento_bancario(valor: Any) -> str:
    """
    Retorna siempre un número para columnas bancarias de documento/factura.

    En Solicitudes de Fondos el campo puede venir como:
      - "FACTURA|12345"
      - "SIN_DOCUMENTO|"
      - "" / None

    Para Santander no debe viajar texto; si no hay número real usamos
    el comodín operativo definido: 999999999.
    """
    raw = _clean_str(valor)
    if "|" in raw:
        raw = raw.split("|", 1)[1]
    numero = re.sub(r"\D", "", raw)
    return numero or "999999999"


def rut_plain_pago_tercero(tercero: Any) -> str | None:
    """
    RUT bancario para pagos a terceros.

    Regla de negocio:
    - Prioriza terceros.rut_cta_num + terceros.rut_cta_dv.
    - Si el tercero no tiene RUT de cuenta informado, usa terceros.rut_num + terceros.dv.
    """
    rut_cta_num = _get_first_attr(tercero, "rut_cta_num", "rut_cuenta_num", "rut_cuenta_bancaria")
    rut_cta_dv = _get_first_attr(tercero, "rut_cta_dv", "dv_cuenta", "dv_cuenta_bancaria")

    rut_num = rut_cta_num if rut_cta_num not in (None, "") else _get_first_attr(tercero, "rut_num", "rut")
    dv = rut_cta_dv if rut_cta_dv not in (None, "") else _get_first_attr(tercero, "dv")

    return rut_plain_from_parts(str(rut_num or ""), dv)


# ============================================================
# Solicitudes de Fondos -> Proveedores (NUEVO)
# ============================================================

def build_proveedor_rows_from_sf_items(
    sf_items: Iterable[Any],
) -> ProveedorBuildResult:
    ok: List[ProveedorRow] = []
    excluidos: List[Dict[str, Any]] = []

    # --- NUEVO: acumulador de REND por rendición ---
    # key: rendicion_id -> dict con totals y datos del usuario
    rend_acc: Dict[int, Dict[str, Any]] = {}

    for it in sf_items:
        item_id = int(getattr(it, "id", 0) or 0)
        seccion = (getattr(it, "seccion", "") or "").strip().upper()
        monto_item = _to_decimal(getattr(it, "monto", None))

        sol = getattr(it, "solicitud", None)
        centro_costo = (getattr(sol, "centro_costo", "") or "").strip() if sol else ""

        if monto_item <= 0:
            excluidos.append(
                {
                    "item_id": item_id,
                    "seccion": seccion,
                    "motivo": "monto <= 0",
                    "rut": None,
                    "nombre": getattr(it, "beneficiario_nombre", None),
                }
            )
            continue

        # ============================================================
        # REND: NO pagamos proveedores; pagamos al USUARIO/TRABAJADOR
        # ============================================================
        if seccion == "REND":
            # intentamos ubicar la rendición relacionada
            rend = _get_first_attr(it, "rendicion", "rendicion_gasto", "rg", "rend")
            rend_id = (
                _get_first_attr(it, "rendicion_id", "rendicion_gasto_id", "rg_id", "rend_id")
                or (getattr(rend, "id", None) if rend is not None else None)
            )

            if not rend_id:
                excluidos.append(
                    {
                        "item_id": item_id,
                        "seccion": seccion,
                        "motivo": "REND sin rendicion_id/relación a rendición",
                        "rut": None,
                        "nombre": getattr(it, "beneficiario_nombre", None),
                    }
                )
                continue

            rend_id_int = int(rend_id)

            # datos del usuario que rindió (defensivo: prueba campos típicos)
            usuario_nombre = (
                _clean_str(_get_first_attr(rend, "usuario_nombre", "usuario", "creado_por_nombre", "creado_por"))
                or _clean_str(_get_first_attr(it, "usuario_nombre", "creado_por_nombre"))
            )

            usuario_rut = (
                _clean_str(_get_first_attr(rend, "usuario_rut", "rut_usuario", "rut", "creado_por_rut"))
                or _clean_str(_get_first_attr(it, "usuario_rut", "rut_usuario", "rut"))
            )

            # Si el rut viene como "13367585-K" o "13367585K", lo normalizamos
            usuario_rut_plain = rut_plain_from_str(usuario_rut)

            # acumulamos monto total de la rendición
            if rend_id_int not in rend_acc:
                rend_acc[rend_id_int] = {
                    "rend_id": rend_id_int,
                    "centro_costo": centro_costo,
                    "monto_total": Decimal("0"),
                    "usuario_nombre": usuario_nombre,
                    "usuario_rut_plain": usuario_rut_plain,
                    # guardamos además el primer item_id para reportes si se excluye
                    "first_item_id": item_id,
                    # intentamos capturar email si existiera
                    "usuario_email": _clean_str(_get_first_attr(rend, "usuario_email", "correo", "email", "creado_por_email")),
                    # si ya viene trabajador linkeado en algún item, lo guardamos (golden path)
                    "trabajador_obj": getattr(it, "trabajador", None),
                }

            rend_acc[rend_id_int]["monto_total"] = rend_acc[rend_id_int]["monto_total"] + monto_item

            # si este item sí trae trabajador y el acumulador aún no lo tiene, lo guardamos
            if rend_acc[rend_id_int].get("trabajador_obj") is None and getattr(it, "trabajador", None) is not None:
                rend_acc[rend_id_int]["trabajador_obj"] = getattr(it, "trabajador", None)

            # seguimos al siguiente item (no generamos fila por ítem)
            continue

        # ============================================================
        # NO-REND (TERC/VAR/etc): flujo actual por ítem (se mantiene)
        # ============================================================
        tercero = getattr(it, "tercero", None)
        trabajador = getattr(it, "trabajador", None)

        if tercero is not None:
            # Para terceros, la nómina bancaria debe usar el RUT de la cuenta
            # bancaria (rut_cta_num + rut_cta_dv), con fallback al RUT principal.
            rut_b = rut_plain_pago_tercero(tercero)
            nombre_b = _clean_str(getattr(tercero, "nombre_display", None))
            banco_code = _clean_str(getattr(getattr(tercero, "banco", None), "codigo_sbif", None))
            cuenta_b = _account_plain(getattr(tercero, "cuenta_numero", None))
            email_b = _email_or_fallback(getattr(tercero, "correo", None))
        elif trabajador is not None:
            rut_b = rut_plain_from_parts(getattr(trabajador, "rut", None), getattr(trabajador, "dv", None))
            nombre_b = " ".join(
                p for p in [
                    _clean_str(getattr(trabajador, "ap_paterno", None)),
                    _clean_str(getattr(trabajador, "ap_materno", None)),
                    _clean_str(getattr(trabajador, "nombres", None)),
                ] if p
            ).strip()
            banco_code = _clean_str(getattr(getattr(trabajador, "banco", None), "codigo_sbif", None))
            cuenta_raw = _clean_str(getattr(trabajador, "cuenta_numero", None)) or _clean_str(getattr(trabajador, "cuenta_rut", None))
            cuenta_b = _account_plain(cuenta_raw)
            email_b = _email_or_fallback(getattr(trabajador, "correo", None))
        else:
            rut_b = rut_plain_from_parts(str(getattr(it, "beneficiario_rut_num", "") or ""), getattr(it, "beneficiario_dv", None))
            nombre_b = _clean_str(getattr(it, "beneficiario_nombre", None))
            banco_code = ""
            cuenta_b = ""
            email_b = _email_or_fallback(None)

        if not rut_b:
            excluidos.append({"item_id": item_id, "seccion": seccion, "motivo": "RUT beneficiario inválido/vacío", "rut": None, "nombre": nombre_b})
            continue
        if not nombre_b:
            excluidos.append({"item_id": item_id, "seccion": seccion, "motivo": "Nombre beneficiario vacío", "rut": rut_b, "nombre": None})
            continue
        if not _is_valid_bank_code(banco_code):
            excluidos.append({"item_id": item_id, "seccion": seccion, "motivo": f"Código banco inválido: {banco_code or '(vacío)'}", "rut": rut_b, "nombre": nombre_b})
            continue
        if not _is_valid_account(cuenta_b):
            excluidos.append({"item_id": item_id, "seccion": seccion, "motivo": "Cuenta destino inválida/vacía", "rut": rut_b, "nombre": nombre_b})
            continue

        nro_fact = _numero_documento_bancario(getattr(it, "nro_documento", None))

        glosa = _clean_str(getattr(it, "descripcion", None)) or _clean_str(getattr(it, "concepto_label", None)) or "Pago"

        ok.append(
            ProveedorRow(
                item_id=item_id,
                seccion=seccion,
                centro_costo=centro_costo,
                rut_beneficiario=rut_b,
                nombre_beneficiario=nombre_b,
                codigo_banco=banco_code,
                cuenta_destino=cuenta_b,
                email_beneficiario=email_b,
                nro_fact_1=nro_fact,
                monto=monto_item,
                glosa=glosa,
            )
        )

    # ============================================================
    # Finalizamos REND: 1 fila por rendición (pago al trabajador)
    # ============================================================
    for rend_id, data in rend_acc.items():
        monto_total = data["monto_total"]
        if monto_total <= 0:
            excluidos.append(
                {
                    "item_id": data.get("first_item_id"),
                    "seccion": "REND",
                    "motivo": "REND monto_total <= 0",
                    "rut": data.get("usuario_rut_plain"),
                    "nombre": data.get("usuario_nombre"),
                }
            )
            continue

        trabajador = data.get("trabajador_obj")

        # Si no vino trabajador linkeado, al menos necesitamos rut del usuario
        if trabajador is None:
            rut_plain = data.get("usuario_rut_plain")
            if not rut_plain:
                excluidos.append(
                    {
                        "item_id": data.get("first_item_id"),
                        "seccion": "REND",
                        "motivo": "REND sin RUT de usuario (no se puede hacer match con Trabajador)",
                        "rut": None,
                        "nombre": data.get("usuario_nombre"),
                    }
                )
                continue

            # ⚠️ Aquí es donde idealmente haces el MATCH real con BD:
            # - Este service no tiene db.session ni modelos importados,
            #   así que NO podemos consultar Trabajador desde acá sin acoplar.
            #
            # Solución recomendada: en routes.py, cuando armes sf_items para generar nómina,
            # JOIN contra Trabajador por rut/dv y setea it.trabajador (o un atributo extra).
            #
            # Mientras no exista ese join, debemos excluir para evitar pagos erróneos.
            excluidos.append(
                {
                    "item_id": data.get("first_item_id"),
                    "seccion": "REND",
                    "motivo": "REND requiere match a Trabajador por RUT (hacer JOIN en query y poblar it.trabajador)",
                    "rut": rut_plain,
                    "nombre": data.get("usuario_nombre"),
                }
            )
            continue

        # Con trabajador ya resuelto: sacamos datos bancarios igual que en caso trabajador normal
        rut_b = rut_plain_from_parts(getattr(trabajador, "rut", None), getattr(trabajador, "dv", None))
        nombre_b = " ".join(
            p for p in [
                _clean_str(getattr(trabajador, "ap_paterno", None)),
                _clean_str(getattr(trabajador, "ap_materno", None)),
                _clean_str(getattr(trabajador, "nombres", None)),
            ] if p
        ).strip()

        banco_code = _clean_str(getattr(getattr(trabajador, "banco", None), "codigo_sbif", None))
        cuenta_raw = _clean_str(getattr(trabajador, "cuenta_numero", None)) or _clean_str(getattr(trabajador, "cuenta_rut", None))
        cuenta_b = _account_plain(cuenta_raw)
        email_b = _email_or_fallback(_get_first_attr(trabajador, "correo", "email", "mail") or data.get("usuario_email"))

        if not rut_b:
            excluidos.append({"item_id": data.get("first_item_id"), "seccion": "REND", "motivo": "REND RUT trabajador inválido/vacío", "rut": None, "nombre": nombre_b})
            continue
        if not nombre_b:
            excluidos.append({"item_id": data.get("first_item_id"), "seccion": "REND", "motivo": "REND Nombre trabajador vacío", "rut": rut_b, "nombre": None})
            continue
        if not _is_valid_bank_code(banco_code):
            excluidos.append({"item_id": data.get("first_item_id"), "seccion": "REND", "motivo": f"REND Código banco inválido: {banco_code or '(vacío)'}", "rut": rut_b, "nombre": nombre_b})
            continue
        if not _is_valid_account(cuenta_b):
            excluidos.append({"item_id": data.get("first_item_id"), "seccion": "REND", "motivo": "REND Cuenta destino inválida/vacía", "rut": rut_b, "nombre": nombre_b})
            continue

        # 1 fila por rendición
        glosa = f"Rendición de gastos - {data.get('centro_costo') or ''}".strip()

        ok.append(
            ProveedorRow(
                item_id=int(data.get("first_item_id") or 0),
                seccion="REND",
                centro_costo=data.get("centro_costo") or "",
                rut_beneficiario=rut_b,
                nombre_beneficiario=nombre_b,
                codigo_banco=banco_code,
                cuenta_destino=cuenta_b,
                email_beneficiario=email_b,
                nro_fact_1="999999999",
                monto=monto_total,
                glosa=glosa,
            )
        )

    return ProveedorBuildResult(ok_rows=ok, excluidos=excluidos)


# ============================================================
# Escritura XLSX desde plantilla (existente + NUEVO PAGOS_PROV)
# ============================================================

def _norm_header(s: str) -> str:
    s = _clean_str(s).lower()
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        m[_norm_header(v)] = col
    return m


def _clear_data_rows(ws, start_row: int = 2) -> None:
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def build_xlsx_from_template(
    template_path: str,
    template_kind: str,  # "PAGOS_REMUN" | "TRANSFER_MAS" | "PAGOS_PROV"
    rows: List[Any],
    anio: int,
    mes: int,
    glosa: Optional[str] = None,
    cuenta_origen_transfer_mas: str = "5555555555",
) -> Workbook:
    wb = load_workbook(template_path)
    ws = wb.active

    headers = _header_map(ws)
    _clear_data_rows(ws, start_row=2)

    glosa_final = glosa if (glosa is not None and str(glosa).strip()) else glosa_anticipo(anio, mes)

    if template_kind == "PAGOS_REMUN":
        def col(name: str) -> int:
            key = _norm_header(name)
            if key not in headers:
                raise ValueError(f"Plantilla inválida: no se encontró columna '{name}'.")
            return headers[key]

        c_rut = col("RUT Beneficiario")
        c_nom = col("Nombre Beneficiario")
        c_monto = col("Monto")
        c_banco = col("Código Banco")
        c_cta = col("Cuenta destino")
        c_mod = col("Modalidad de pago")
        c_suc = col("Sucursal entrega")
        c_mail = col("Mail beneficiario")
        c_glosa = col("Glosa mail")

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=c_rut).value = r.rut_beneficiario
            ws.cell(row=i, column=c_nom).value = r.nombre_beneficiario
            ws.cell(row=i, column=c_monto).value = float(r.monto)
            ws.cell(row=i, column=c_banco).value = r.codigo_banco
            ws.cell(row=i, column=c_cta).value = r.cuenta_destino
            ws.cell(row=i, column=c_mod).value = "3"
            ws.cell(row=i, column=c_suc).value = "999"
            ws.cell(row=i, column=c_mail).value = r.email_beneficiario or "transferencias@grupo-cs.cl"
            glosa_row = getattr(r, "glosa", None) or glosa_final
            ws.cell(row=i, column=c_glosa).value = glosa_row

        return wb

    if template_kind == "TRANSFER_MAS":
        def col(name: str) -> int:
            key = _norm_header(name)
            if key not in headers:
                raise ValueError(f"Plantilla inválida: no se encontró columna '{name}'.")
            return headers[key]

        c_origen = col("Cuenta Origen")
        c_mon1 = col("Moneda")
        c_cta = col("Cuenta destino")
        c_mon2 = col("Moneda2")
        c_banco = col("Codigo Banco")
        c_rut = col("Rut Beneficiario")
        c_nom = col("Nombre Beneficiario")
        c_monto = col("Monto")
        c_glosa = col("Glosa")
        c_mail = col("Direccion correo")
        c_glosa_mail = col("Glosa correo")
        c_cart_cli = col("Glosa cartola cliente")
        c_cart_ben = col("Glosa Cartola Beneficiario")

        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=c_origen).value = cuenta_origen_transfer_mas
            ws.cell(row=i, column=c_mon1).value = "CLP"
            ws.cell(row=i, column=c_cta).value = r.cuenta_destino
            ws.cell(row=i, column=c_mon2).value = "CLP"
            ws.cell(row=i, column=c_banco).value = r.codigo_banco
            ws.cell(row=i, column=c_rut).value = r.rut_beneficiario
            ws.cell(row=i, column=c_nom).value = r.nombre_beneficiario
            ws.cell(row=i, column=c_monto).value = float(r.monto)
            ws.cell(row=i, column=c_mail).value = r.email_beneficiario or "transferencias@grupo-cs.cl"            
            glosa_row = getattr(r, "glosa", None) or glosa_final
            ws.cell(row=i, column=c_glosa).value = glosa_row
            ws.cell(row=i, column=c_glosa_mail).value = glosa_row
            ws.cell(row=i, column=c_cart_cli).value = glosa_row
            ws.cell(row=i, column=c_cart_ben).value = glosa_row

        return wb

    if template_kind == "PAGOS_PROV":
        def col_any(*names: str) -> int:
            for name in names:
                key = _norm_header(name)
                if key in headers:
                    return headers[key]
            raise ValueError(f"Plantilla inválida: no se encontró ninguna de las columnas {names}.")

        c_rut = col_any("Rut Beneficiario", "RUT Beneficiario", "RUT Proveedor", "Rut Proveedor")
        c_nom = col_any("Nombre Beneficiario", "Nombre Proveedor")
        c_mod = col_any("Modalidad Pago", "Modalidad pago", "Modalidad de pago")
        c_banco = col_any("Código Banco", "Codigo Banco", "Banco")
        c_cta = col_any("Cuenta destino", "Nro. Cuenta", "Nro Cuenta")
        c_suc = col_any("Cód. Sucursal", "Cod. Sucursal", "Codigo Sucursal")
        c_monto_total = col_any("Monto Total", "Monto total")
        c_mail = col_any("Direccion correo", "Dirección correo", "Correo", "Mail")
        c_glosa = col_any("Glosa", "Glosa ")

        # columnas facturas 1..11
        fact_cols: List[Tuple[int, int]] = []
        for i in range(1, 12):
            c_f = col_any(f"Nro. Fact. {i}", f"Nro Fact. {i}")
            c_m = col_any(f"Monto Fact. {i}", f"Monto Fact {i}", f"Monto Fact. {i}")
            fact_cols.append((c_f, c_m))

        if not rows:
            return wb

        for irow, r in enumerate(rows, start=2):
            ws.cell(row=irow, column=c_rut).value = r.rut_beneficiario
            ws.cell(row=irow, column=c_nom).value = r.nombre_beneficiario
            ws.cell(row=irow, column=c_mod).value = "3"
            ws.cell(row=irow, column=c_banco).value = r.codigo_banco
            ws.cell(row=irow, column=c_cta).value = r.cuenta_destino
            ws.cell(row=irow, column=c_suc).value = "999"

            # limpiar 11 slots (por si la plantilla trae residuos)
            for (cf, cm) in fact_cols:
                ws.cell(row=irow, column=cf).value = None
                ws.cell(row=irow, column=cm).value = None

                        # poblar facturas
            total = Decimal("0")

            # 1) Caso ProveedorPagoRow (pago_proveedores): trae lista facturas[]
            facturas = getattr(r, "facturas", None)

            # 2) Caso ProveedorRow (solicitudes_fondos): NO trae facturas[]
            #    -> usamos nro_fact_1 + monto como Factura 1
            if not facturas:
                nro1 = _clean_str(getattr(r, "nro_fact_1", None))
                monto1 = _to_decimal(getattr(r, "monto", None))

                # Si no hay nro_fact_1, por compatibilidad ponemos el dummy bancario
                if monto1 > 0:
                    facturas = [{"nro": _numero_documento_bancario(nro1), "monto": monto1}]
                else:
                    facturas = []

            for idx, f in enumerate(facturas[:11]):
                nro = _clean_str(f.get("nro"))
                monto = _to_decimal(f.get("monto"))
                cf, cm = fact_cols[idx]
                ws.cell(row=irow, column=cf).value = _numero_documento_bancario(nro)
                ws.cell(row=irow, column=cm).value = float(monto)
                total += monto

            # Total: preferimos monto_total (ProveedorPagoRow), luego monto (ProveedorRow), luego suma calculada
            monto_total = getattr(r, "monto_total", None)
            if monto_total is None or _to_decimal(monto_total) <= 0:
                monto_total = _to_decimal(getattr(r, "monto", None))
            if monto_total is None or _to_decimal(monto_total) <= 0:
                monto_total = total

            ws.cell(row=irow, column=c_monto_total).value = float(_to_decimal(monto_total))
            ws.cell(row=irow, column=c_mail).value = r.email_beneficiario or "transferencias@grupo-cs.cl"
            ws.cell(row=irow, column=c_glosa).value = _clean_str(getattr(r, "glosa", "")) or glosa_final

        return wb

    raise ValueError(f"template_kind inválido: {template_kind}")


def get_template_path(app_root_path: str, template_filename: str) -> str:
    return os.path.join(app_root_path, "static", "xlsx_templates", template_filename)