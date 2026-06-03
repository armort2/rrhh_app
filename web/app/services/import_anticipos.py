from __future__ import annotations

import re
from decimal import Decimal
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func

from ..extensions import db
from ..models import Trabajador, Obra, Contrato, AnticipoNomina, AnticipoDetalle


_RUT_CLEAN_RE = re.compile(r"[^0-9kK]")


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip().lower()


def split_rut(raw: Any) -> tuple[str, str]:
    """
    Convierte un RUT en formato libre (12.345.678-9 / 12345678-9 / 123456789)
    a (rut_digits, dv). Retorna ("","") si inválido.
    """
    if raw is None:
        return "", ""

    cleaned = _RUT_CLEAN_RE.sub("", str(raw)).upper()
    if len(cleaned) < 2:
        return "", ""

    dv = cleaned[-1]
    rut_digits = re.sub(r"\D", "", cleaned[:-1])

    if not rut_digits or not dv:
        return "", ""

    if not (dv.isdigit() or dv == "K"):
        return "", ""

    return rut_digits, dv


def _clean_rut_compact(raw: Any) -> str | None:
    """
    Devuelve RUT compactado (solo dígitos + DV) para deduplicación.
    Ej: 12.345.678-9 -> 123456789
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    return _RUT_CLEAN_RE.sub("", s).upper() or None


def _parse_money(v: Any) -> Decimal:
    if v is None:
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))

    s = str(v).strip()
    if not s:
        return Decimal("0")

    s = s.replace("$", "").replace("CLP", "").replace(" ", "")

    # Si trae formato con miles: 120.000 -> 120000
    # Si trae ambos "." y "," asumimos "." miles y "," decimal
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "").replace(",", "")

    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


_OBRA_PREFIX_RE = re.compile(r"^\s*obra\s+", re.IGNORECASE)


def _normalize_obra_name_for_lookup(name: str) -> str:
    """
    Normalización conservadora para lookup de obras:
    - trim
    - colapsa espacios
    - elimina prefijo "Obra " si existe
    """
    s = re.sub(r"\s+", " ", (name or "")).strip()
    s = _OBRA_PREFIX_RE.sub("", s).strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_obra_by_excel_name(obra_excel_name: str | None) -> Obra | None:
    """
    Estrategia robusta:
    1) match exacto por nombre (tal cual)
    2) match por nombre sin prefijo "Obra "
    3) match por nombre colapsando espacios (sin prefijo)
    """
    if not obra_excel_name:
        return None

    raw = str(obra_excel_name).strip()
    if not raw:
        return None

    # 1) exacto tal cual viene
    obra = Obra.query.filter_by(nombre=raw).first()
    if obra:
        return obra

    # 2) sin prefijo "Obra "
    no_prefix = _OBRA_PREFIX_RE.sub("", raw).strip()
    if no_prefix and no_prefix != raw:
        obra = Obra.query.filter_by(nombre=no_prefix).first()
        if obra:
            return obra

    # 3) colapsar espacios y sin prefijo (más tolerante, pero aún conservador)
    normalized = _normalize_obra_name_for_lookup(raw)
    if normalized and normalized not in {raw, no_prefix}:
        obra = Obra.query.filter_by(nombre=normalized).first()
        if obra:
            return obra

    return None


def _contratos_trabajador_periodo_cc(
    trabajador_id: int,
    *,
    anio: int,
    mes: int,
    centro_costo: str,
    obra_id: int | None = None,
) -> list[Contrato]:
    """
    Contratos del trabajador vigentes/intersectados en el período de anticipos,
    filtrados por centro de costo de la obra. Si Excel trae obra y esta se logró
    resolver, se prefiere esa obra para evitar tomar contratos de otra obra.
    """
    periodo_inicio = date(int(anio), int(mes), 1)
    periodo_fin = date(int(anio) + 1, 1, 1) if int(mes) == 12 else date(int(anio), int(mes) + 1, 1)

    q = (
        Contrato.query
        .join(Obra, Obra.id == Contrato.obra_id)
        .filter(Contrato.trabajador_id == trabajador_id)
        .filter(func.upper(func.coalesce(Obra.centro_costo, "")) == (centro_costo or "").strip().upper())
        .filter(Contrato.fecha_inicio.isnot(None))
        .filter(Contrato.fecha_inicio < periodo_fin)
        .filter(func.coalesce(Contrato.fecha_termino, date(2999, 12, 31)) >= periodo_inicio)
    )

    if obra_id:
        q = q.filter(Contrato.obra_id == obra_id)

    return q.order_by(Contrato.id.desc()).all()


def import_anticipos_from_excel(
    *,
    filepath: str,
    anio: int,
    mes: int,
    centro_costo: str,
    user_id: int | None = None,
    allow_reimport: bool = True,
) -> dict[str, Any]:
    """
    Importa anticipos desde Excel.
    - Hoja: 'Anticipos'
    - Monto: 'Anticipo mes actual'
    - Visación: por Centro de Costo (cabecera)
    - Matching de obra: exacto por obras.nombre (con fallback eliminando prefijo "Obra ")
    - Matching de trabajador: trabajadores.rut (dígitos) + trabajadores.dv
    """
    cc = (centro_costo or "").strip().upper()
    if not cc:
        raise ValueError("Centro de costo es obligatorio.")

    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"No existe archivo: {filepath}")

    wb = load_workbook(str(path), data_only=True, read_only=True)
    if "Anticipos" not in wb.sheetnames:
        raise ValueError("El archivo no contiene la hoja 'Anticipos'.")

    ws = wb["Anticipos"]

    # Encabezados (fila 1)
    header_row = 1
    headers: dict[str, int] = {}
    for col in range(1, 60):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue
        headers[_norm(str(val))] = col

    def col_of(*names: str) -> int | None:
        for n in names:
            k = _norm(n)
            if k in headers:
                return headers[k]
        return None

    col_rut = col_of("RUT")
    col_monto = col_of("Anticipo mes actual")
    col_obra = col_of("Nombre Obra")
    col_nombre = col_of("Apellidos/Nombres", "Nombre")
    col_obs = col_of("Observaciones", "Observación")
    col_cc = col_of("CENTRO DE COSTO", "Centro de costo")

    if not col_rut or not col_monto:
        raise ValueError("Faltan columnas mínimas: 'RUT' y/o 'Anticipo mes actual'.")

    if not col_obra:
        raise ValueError("Falta columna 'Nombre Obra' (necesaria para validar CC y trazabilidad).")

    # Nómina: crear o cargar
    nomina = AnticipoNomina.query.filter_by(anio=anio, mes=mes, centro_costo=cc).first()
    if nomina and not allow_reimport:
        raise ValueError("Ya existe nómina para ese período y centro de costo.")
    if not nomina:
        nomina = AnticipoNomina(anio=anio, mes=mes, centro_costo=cc)
        db.session.add(nomina)
        db.session.flush()
    else:
        AnticipoDetalle.query.filter_by(nomina_id=nomina.id).delete()

    nomina.estado = "BORRADOR"
    nomina.origen_archivo = path.name
    nomina.importado_por = user_id
    nomina.importado_en = datetime.utcnow()

    stats = {
        "filas": 0,
        "ok": 0,
        "no_encontrado": 0,
        "obra_no_calza": 0,
        "duplicadas": 0,
        "observadas": 0,
    }

    seen_ruts: set[str] = set()

    for row in range(header_row + 1, ws.max_row + 1):
        rut_raw = ws.cell(row=row, column=col_rut).value
        monto_raw = ws.cell(row=row, column=col_monto).value
        obra_raw = ws.cell(row=row, column=col_obra).value

        if rut_raw is None and monto_raw is None and obra_raw is None:
            continue

        rut_compact = _clean_rut_compact(rut_raw)
        rut_digits, rut_dv = split_rut(rut_raw)

        monto = _parse_money(monto_raw)
        obra_nombre_excel = str(obra_raw).strip() if obra_raw else ""

        if not rut_compact and monto == 0 and not obra_nombre_excel:
            continue

        stats["filas"] += 1

        estado_linea = "OK"
        obs = ws.cell(row=row, column=col_obs).value if col_obs else None
        obs_text = str(obs).strip() if obs else None

        # Validación centro costo (si viene en excel)
        if col_cc:
            cc_excel = ws.cell(row=row, column=col_cc).value
            if cc_excel:
                cc_excel_norm = str(cc_excel).strip().upper()
                if cc_excel_norm and cc_excel_norm != cc:
                    if estado_linea == "OK":
                        estado_linea = "OBSERVADA"
                        stats["observadas"] += 1
                    obs_text = (obs_text + " | " if obs_text else "") + (
                        f"Centro costo Excel='{cc_excel_norm}' distinto a '{cc}'"
                    )

        # Duplicado por RUT compactado
        if rut_compact:
            if rut_compact in seen_ruts:
                estado_linea = "DUPLICADA"
                stats["duplicadas"] += 1
            else:
                seen_ruts.add(rut_compact)

        # Obra (robusto a prefijo "Obra ")
        obra = _find_obra_by_excel_name(obra_nombre_excel) if obra_nombre_excel else None

        # Validar obra pertenece al CC de la nómina (si fue encontrada)
        if obra and (obra.centro_costo or "").strip().upper() != cc:
            estado_linea = "OBRA_NO_CALZA"
            stats["obra_no_calza"] += 1

        # Trabajador por RUT + DV
        trabajador = None
        if rut_digits and rut_dv:
            trabajador = (
                Trabajador.query
                .filter(Trabajador.rut == rut_digits)
                .filter(Trabajador.dv == rut_dv)
                .first()
            )

        if not trabajador:
            if estado_linea == "OK":
                estado_linea = "NO_ENCONTRADO"
            stats["no_encontrado"] += 1

        # Snapshot nombre
        nombre_excel = ws.cell(row=row, column=col_nombre).value if col_nombre else None
        nombre_snapshot = str(nombre_excel).strip() if nombre_excel else None
        if not nombre_snapshot and trabajador:
            parts = [trabajador.nombres, trabajador.ap_paterno, trabajador.ap_materno]
            nombre_snapshot = " ".join([p for p in parts if p])

        # Snapshot rut legible
        rut_snapshot = f"{rut_digits}-{rut_dv}" if rut_digits and rut_dv else (str(rut_raw).strip() if rut_raw else None)

        contratos = []
        if trabajador:
            contratos = _contratos_trabajador_periodo_cc(
                trabajador.id,
                anio=anio,
                mes=mes,
                centro_costo=cc,
                obra_id=obra.id if obra else None,
            )

            if not contratos:
                if estado_linea == "OK":
                    estado_linea = "SIN_CONTRATO_EN_PERIODO"
                    obs_text = (obs_text + " | " if obs_text else "") + "Sin contrato vigente en período/CC"

        # Regla conservadora:
        # - Si hay 1 contrato, la línea queda contract-aware.
        # - Si hay varios contratos para el mismo trabajador/obra/CC, se crea una línea por contrato.
        #   El monto importado queda en el contrato más reciente y las demás líneas quedan en $0
        #   para que RRHH distribuya el anticipo sin duplicar pago.
        contratos_para_insertar = contratos or [None]

        for idx, contrato in enumerate(contratos_para_insertar):
            estado_insert = estado_linea
            obs_insert = obs_text
            monto_insert = monto

            if trabajador and len(contratos_para_insertar) > 1:
                estado_insert = "AMBIGUO_CONTRATO"
                obs_insert = (obs_insert + " | " if obs_insert else "") + (
                    f"Trabajador con {len(contratos_para_insertar)} contratos en período/CC; "
                    "revisar distribución del anticipo por contrato"
                )
                if idx > 0:
                    monto_insert = Decimal("0")

            detalle = AnticipoDetalle(
                nomina_id=nomina.id,
                trabajador_id=trabajador.id if trabajador else None,
                contrato_id=contrato.id if contrato else None,
                obra_id=(contrato.obra_id if contrato and getattr(contrato, "obra_id", None) else (obra.id if obra else None)),
                rut=rut_snapshot,
                nombre_completo=nombre_snapshot,
                nombre_obra=(getattr(getattr(contrato, "obra", None), "nombre", None) if contrato else obra_nombre_excel) or obra_nombre_excel,
                centro_costo=cc,
                monto=monto_insert,
                observacion=obs_insert,
                estado_linea=estado_insert,
            )
            db.session.add(detalle)

            if estado_insert == "OK":
                stats["ok"] += 1
            elif estado_insert == "AMBIGUO_CONTRATO":
                stats["observadas"] += 1

    db.session.commit()
    return {
        "nomina_id": nomina.id,
        "periodo": f"{anio}-{mes:02d}",
        "centro_costo": cc,
        "stats": stats,
    }
