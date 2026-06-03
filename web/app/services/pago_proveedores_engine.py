# web/app/services/pago_proveedores_engine.py
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook

from .bank_nominas import (
    rut_plain_from_str,
    rut_plain_from_parts,
    _to_decimal,
    _clean_str,
    _account_plain,
    _email_or_fallback,
    _is_valid_account,
    _is_valid_bank_code,
    ProveedorPagoRow,
)

# -------------------------
# Helpers RUT
# -------------------------
def rut_parts_from_plain(rut_plain: str) -> Tuple[int | None, str | None]:
    """
    rut_plain: '76123456K' -> (76123456, 'K')
    """
    s = _clean_str(rut_plain).upper()
    s = re.sub(r"[^0-9K]", "", s)
    if len(s) < 2:
        return None, None
    base = s[:-1]
    dv = s[-1]
    if not base.isdigit():
        return None, None
    return int(base), dv


def norm_pagador(p: str) -> str:
    s = _clean_str(p).upper()
    if "VALEN" in s:
        return "VALENCIA"
    if "SALEM" in s:
        return "SALEM"
    return s or "PAGADOR"


@dataclass(frozen=True)
class ParsedLinea:
    pagador: str
    rut_tercero_plain: str
    nro_fact: str
    monto: Decimal


@dataclass(frozen=True)
class EngineResult:
    ok_by_pagador: Dict[str, List[ProveedorPagoRow]]
    excluidos: List[Dict[str, Any]]


def parse_input_xlsx(path: str) -> Tuple[List[ParsedLinea], List[Dict[str, Any]]]:
    wb = load_workbook(path)
    ws = wb.active

    # headers esperados (tu archivo trae: Pagador, RUT, Correl, Nro. Fact., Monto)
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[_clean_str(v).strip().lower()] = c

    def col(*names: str) -> int | None:
        for n in names:
            k = n.lower()
            if k in headers:
                return headers[k]
        return None

    c_pag = col("pagador")
    c_rut = col("rut")
    c_fact = col("nro. fact.", "nro fact.", "nro factura", "factura")
    c_monto = col("monto")

    excl: List[Dict[str, Any]] = []
    out: List[ParsedLinea] = []

    if not all([c_pag, c_rut, c_fact, c_monto]):
        excl.append({"motivo": "Plantilla de carga inválida: faltan columnas (Pagador, RUT, Nro. Fact., Monto)"})
        return out, excl

    for r in range(2, ws.max_row + 1):
        pag = _clean_str(ws.cell(r, c_pag).value)
        rut_in = rut_plain_from_str(_clean_str(ws.cell(r, c_rut).value))
        fact = _clean_str(ws.cell(r, c_fact).value)
        monto = _to_decimal(ws.cell(r, c_monto).value)

        # saltar filas vacías “de relleno”
        if not pag and not rut_in and not fact and monto == 0:
            continue

        if not rut_in:
            excl.append({"fila": r, "motivo": "RUT vacío/invalidable", "pagador": pag, "rut": ws.cell(r, c_rut).value})
            continue
        if not fact:
            excl.append({"fila": r, "motivo": "Nro. Fact. vacío", "pagador": pag, "rut": rut_in})
            continue
        if monto <= 0:
            excl.append({"fila": r, "motivo": "Monto <= 0", "pagador": pag, "rut": rut_in, "fact": fact})
            continue

        out.append(
            ParsedLinea(
                pagador=norm_pagador(pag),
                rut_tercero_plain=rut_in,
                nro_fact=fact,
                monto=monto,
            )
        )

    return out, excl


def build_rows(
    lineas: List[ParsedLinea],
    terceros_by_rut: Dict[Tuple[int, str], Any],  # (rut_num, dv) -> Tercero SQLAlchemy
) -> EngineResult:
    excluidos: List[Dict[str, Any]] = []
    ok_by_pagador: Dict[str, List[ProveedorPagoRow]] = {}

    # agrupación por (pagador, rut_tercero_plain) preservando orden de aparición
    groups: Dict[Tuple[str, str], List[ParsedLinea]] = {}
    order_keys: List[Tuple[str, str]] = []

    for ln in lineas:
        k = (ln.pagador, ln.rut_tercero_plain)
        if k not in groups:
            groups[k] = []
            order_keys.append(k)
        groups[k].append(ln)

    for pag_rut in order_keys:
        pagador, rut_plain = pag_rut
        rut_num, dv = rut_parts_from_plain(rut_plain)
        if rut_num is None or dv is None:
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "RUT no parseable a (num,dv)"})
            continue

        tercero = terceros_by_rut.get((rut_num, dv))
        if tercero is None:
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "Tercero no encontrado en BD"})
            continue

        if _clean_str(getattr(tercero, "estado", "")).upper() != "VIGENTE":
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "Tercero no vigente", "tercero_id": getattr(tercero, "id", None)})
            continue

        # RUT beneficiario: rut_cta_num + rut_cta_dv (fallback a rut_num/dv)
        rut_cta_num = getattr(tercero, "rut_cta_num", None) or getattr(tercero, "rut_num", None)
        rut_cta_dv = getattr(tercero, "rut_cta_dv", None) or getattr(tercero, "dv", None)
        rut_benef = rut_plain_from_parts(str(rut_cta_num or ""), rut_cta_dv)

        nombre = " ".join(
            p for p in [
                _clean_str(getattr(tercero, "ap_paterno", None)),
                _clean_str(getattr(tercero, "ap_materno", None)),
                _clean_str(getattr(tercero, "nombres", None)),
            ] if p
        ).strip()

        # banco: preferimos SBIF desde relación bancos.codigo_sbif; fallback a banco_id
        banco_rel = getattr(tercero, "banco", None)
        codigo_banco = _clean_str(getattr(banco_rel, "codigo_sbif", None)) or _clean_str(getattr(tercero, "banco_id", None))

        cuenta = _account_plain(getattr(tercero, "cuenta_numero", None))
        correo = _email_or_fallback(getattr(tercero, "correo", None))

        if not rut_benef:
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "RUT beneficiario (cta) vacío/invalidable", "tercero_id": getattr(tercero, "id", None)})
            continue
        if not nombre:
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "Nombre beneficiario vacío", "tercero_id": getattr(tercero, "id", None)})
            continue
        if not _is_valid_bank_code(codigo_banco):
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": f"Código banco inválido: {codigo_banco or '(vacío)'}", "tercero_id": getattr(tercero, "id", None)})
            continue
        if not _is_valid_account(cuenta):
            excluidos.append({"pagador": pagador, "rut": rut_plain, "motivo": "Cuenta destino inválida/vacía", "tercero_id": getattr(tercero, "id", None)})
            continue

        facturas = [{"nro": ln.nro_fact, "monto": ln.monto} for ln in groups[pag_rut]]

        # split 11 por fila
        chunks = [facturas[i:i+11] for i in range(0, len(facturas), 11)]
        out_rows: List[ProveedorPagoRow] = []

        for idx, chunk in enumerate(chunks, start=1):
            total = sum((_to_decimal(x["monto"]) for x in chunk), Decimal("0"))
            glosa = "Pago de Factura: " + " - ".join(_clean_str(x["nro"]) for x in chunk if _clean_str(x["nro"]))
            out_rows.append(
                ProveedorPagoRow(
                    pagador=pagador,
                    tercero_id=getattr(tercero, "id", None),
                    rut_tercero=rut_plain,
                    rut_beneficiario=rut_benef,
                    nombre_beneficiario=nombre,
                    codigo_banco=codigo_banco,
                    cuenta_destino=cuenta,
                    email_beneficiario=correo,
                    facturas=chunk,
                    monto_total=total,
                    glosa=glosa,
                    fila_orden=idx,
                )
            )

        ok_by_pagador.setdefault(pagador, []).extend(out_rows)

    return EngineResult(ok_by_pagador=ok_by_pagador, excluidos=excluidos)